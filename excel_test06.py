import openpyxl
import re
import sys
import xlsxwriter
from openpyxl.styles import PatternFill


def calculate_outline(routine_name, routine_level):
    outline_left = number_of_spaces * int(routine_level)
    outline_right = number_char_divider - (len(routine_name) + outline_left)
    return outline_left, outline_right


def renumber_testcase_action(list_action, action_type):

    outline_left, outline_right = calculate_outline(
        dict_routines["renumber_testcase_action"][0], dict_routines["renumber_testcase_action"][1])
    print(char_routine_start * outline_left,
          dict_routines["renumber_testcase_action"][0], char_routine_start * outline_right)
    print(char_debug_print * outline_left, "renumber_testcase_action IN parameter list_action ", list_action)
    print(char_debug_print * outline_left, "renumber_testcase_action IN parameter action_type ", action_type)
    letter_counter = 0
    flag_action_number = ""
    count_action_number = 100
    flag_except = 0
    no_action = 0
    c_steps = 0
    c_results = 0
    temp_list_action = []
    list_numbers = []
    if action_type == "STEPS":
        print(char_debug_print * outline_left, "ACTION = STEPS")
    if action_type == "RESULTS":
        print(char_debug_print * outline_left, "ACTION = RESULTS")
    list_alphabet = ["a.", "b.", "c.", "d.", "f.", "g.", "h.", "i.", "j.", "k.", "l.", "m.", "n.", "o.",
                     "p.", "q.", "r.", "s.", "t.", "u.", "v.", "w.", "x.", "y.", "z."]
    for i in range(0, len(list_action)):
        if "Precondition" in list_action[i]:
            continue
        print(char_debug_print * outline_left, "Action", "i", i, "len(list_action)", len(list_action), list_action[i])
        point_pos = list_action[i].find(".")
        print(char_debug_print * outline_left, "point_pos", point_pos)
        if (point_pos == -1) and (len(list_action[i].strip()) == 0):  # skip blank lines
            print(char_debug_print * outline_left,
                  "skip blank line <---------------------------", "len(list_action[i])", len(list_action[i].strip()))
            continue
        # Take first point after number, not ending point)
        if ("." in list_action[i]) and ((point_pos > 0) and (point_pos < 4)):
            split_number_action = list_action[i].split('.')  # Check first point
            action_number = split_number_action[0]
            temp_number = split_number_action[0]
            letter_counter = 0
            no_action = 1
            flag_action_number = action_number
            print(char_debug_print * outline_left, "Action number", action_number, split_number_action)
        else:
            print(char_debug_print * outline_left, "blabla", list_action[i], "no_action", no_action,
                  "len(list_action[i].strip()", len(list_action[i].strip()))
            if (no_action == 0) and (len(list_action[i].strip()) == 0):  # skip first blank step line
                print(char_debug_print * outline_left, "skip first blank step line")
                continue
            if len(list_action[i].strip()) == 0:  # Blanks
                if flag_action_number != action_number:
                    flag_action_number = action_number
                    letter_counter = 0
                    action_number_next, temp_number = action_number + list_alphabet[letter_counter]
                else:
                    action_number_next, temp_number = action_number + list_alphabet[letter_counter]
                    letter_counter += 1
                    # list_steps[i] = step_number_next + list_steps[i]
                    print(char_debug_print * outline_left, "Steps Blank", list_action[i])
            else:  # Starts with letter, no step number
                regex_action = re.compile(r"^[a-zA-Z][a-zA-Z]", re.MULTILINE)
                list_char_detect = regex_action.findall(list_action[i])
                if (len(list_char_detect) > 0):
                    try:
                        dummy = len(action_number)
                    except UnboundLocalError:
                        print(char_debug_print * outline_left, "EXCEPT")
                        flag_except = 1
                        count_action_number -= 1
                        action_number, temp_number = str(count_action_number)
                        print(char_debug_print * outline_left, "action_number", action_number)
                        pass
                    if flag_action_number != action_number:
                        print(char_debug_print * outline_left, "THEN")
                        flag_action_number = action_number
                        letter_counter = 0
                        action_number_next = action_number + list_alphabet[letter_counter]
                        temp_number = action_number + list_alphabet[letter_counter]
                        if (action_number == "99") and (flag_except == 1):
                            print(char_debug_print * outline_left,
                                  "action_number == 99", action_number, "flag_except", flag_except)
                            action_number_next, temp_number = "99. "
                            flag_except = 0
                        else:
                            count_action_number -= 1
                            action_number = str(count_action_number)
                            print(char_debug_print * outline_left,
                                  "action_number == XX", action_number, "flag_except", flag_except)
                            action_number_next = action_number + ". "
                            temp_number = action_number + ". "
                            flag_except = 0
                        print(char_debug_print * outline_left, "action_number_next", action_number_next, i)
                    else:
                        print(char_debug_print * outline_left, "ELSE")
                        action_number_next = action_number + list_alphabet[letter_counter]
                        temp_number = action_number + list_alphabet[letter_counter]
                        letter_counter += 1
                    list_action[i] = action_number_next + list_action[i]
                    print(char_debug_print * outline_left, "Action char", list_action[i])
                    action_number_next = ""
        print(char_debug_print * outline_left, "i", i, "list_action[i]", list_action[i])
        temp_list_action.append(list_action[i])
        if action_type == "STEPS":
            print(char_debug_print * outline_left, "c_steps", c_steps, "c_results", c_results)
            list_numbers.append(temp_number)
            c_steps += 1
        if action_type == "RESULTS":
            print(char_debug_print * outline_left, "ACTION = RESULTS")
            list_numbers.append(temp_number)
            c_results += 1
    list_action.clear()
    list_action = list(temp_list_action)
    if action_type == "STEPS":
        print(char_debug_print * outline_left, "XX step numbers XX")
        print(char_debug_print * outline_left, list_numbers)
    if action_type == "RESULTS":
        print(char_debug_print * outline_left, "XX result numbers XX")
        print(char_debug_print * outline_left, list_numbers)
    print(char_debug_print * outline_left, "renumber_testcase_action OUT parameter list_action ", list_action)
    print(char_debug_print * outline_left, "renumber_testcase_action OUT parameter list_numbers ", list_numbers)
    print(char_routine_stop * outline_left,
          dict_routines["renumber_testcase_action"][0], char_routine_stop * outline_right)
    return list_action, list_numbers


def select_excel_tab():

    outline_left, outline_right = calculate_outline(
        dict_routines["select_excel_tab"][0], dict_routines["select_excel_tab"][1])
    print(char_routine_start * outline_left, dict_routines["select_excel_tab"][0], char_routine_start * outline_right)
    print(char_debug_print * outline_left, "Number of arguments: ", len(sys.argv), "arguments")
    if len(sys.argv) > 1:
        print(char_debug_print * outline_left, "Argument list: ", str(sys.argv), type(sys.argv))
        if len(sys.argv) == 2:
            if sys.argv[1].isnumeric():
                print(char_debug_print * outline_left, "isnumeric", sys.argv[1])
                number_arg = int(sys.argv[1])
                if (number_arg > 0) and (number_arg <= 9):
                    number_arg -= 1
                    print(" " * (number_of_spaces *
                                 int(dict_routines["select_excel_tab"][1])), number_arg, list_TC_topics[number_arg])
        else:
            for number_arg in range(1, len(sys.argv)):
                print(" " * (number_of_spaces *
                             int(dict_routines["select_excel_tab"][1])), number_arg, list_TC_topics[number_arg])
    print(char_debug_print * outline_left, "select_excel_tab OUT parameter number_arg ", number_arg)
    print(char_routine_stop * outline_left, dict_routines["select_excel_tab"][0], char_routine_stop * outline_right)
    return number_arg


def write_header():

    outline_left, outline_right = calculate_outline(
        dict_routines["write_header"][0], dict_routines["write_header"][1])
    print(char_routine_start * outline_left, dict_routines["write_header"][0], char_routine_start * outline_right)
    #  Print header 1
    header1_row = 0
    header1_col1 = 0
    header1_col2 = 2
    # Create a format to use in the merged range.
    merge_format = wb_template.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '808080'})
    # Merge 2 cells.
    print(char_debug_print * outline_left, "Header1", header1_row,
          header1_col1, list_header1[0])  # Header 1 Column Test Suite
    worksheet.write(header1_row, header1_col1, list_header1[0])
    worksheet.merge_range('A1:B1', list_header1[0], merge_format)
    # Merge 9 cells.
    print(char_debug_print * outline_left, "Header1", header1_row,
          header1_col2, list_header1[1])  # Header 1 Column Test Cases
    worksheet.write(header1_row, header1_col2, list_header1[1])
    worksheet.merge_range('C1:K1', list_header1[1], merge_format)  # format gray text
    #  print header 2
    header2_row = 1
    header2_max_col = 11
    for header2_col in range(header2_max_col):
        print(char_debug_print * outline_left, "Header2", header2_row, header2_col, list_header2[header2_col])
        #    worksheet.cell(header2_row, header2_col).fill = PatternFill(fgColor='808080', fill_type='solid')
        worksheet.write(header2_row, header2_col, list_header2[header2_col], merge_format)
        #    my_fill = PatternFill(fgColor='FFFFFF', bgColor='FFFFFF', fill_type='solid')
        #    worksheet['A2'].fill = my_fill
    print(char_routine_stop * outline_left, dict_routines["write_header"][0], char_routine_stop * outline_right)


def write_to_excel_template(step, result, row, flag_rn, flag_dm):

    outline_left, outline_right = calculate_outline(
        dict_routines["write_to_excel_template"][0], dict_routines["write_to_excel_template"][1])
    print(char_routine_start * outline_left,
          dict_routines["write_to_excel_template"][0], char_routine_start * outline_right)
    print(char_debug_print * outline_left, "write_to_excel_template IN parameter step ", step)
    print(char_debug_print * outline_left, "write_to_excel_template IN parameter result ", result)
    print(char_debug_print * outline_left, "write_to_excel_template IN parameter row ", row)
    print(char_debug_print * outline_left, "write_to_excel_template IN parameter flag_rn ", flag_rn)
    print(char_debug_print * outline_left, "write_to_excel_template IN parameter flag_dm ", flag_dm)

    col = 7  # Steps : read column F(6); Steps : write column H(8)
    if flag_remove_number == 1:
        step = remove_numbers(step)
    worksheet.write(row, col, step)
    col = 8  # Steps : read column F(6); Steps : write column I(9)
    if flag_remove_number == 1:
        result = remove_numbers(result)
    worksheet.write(row, col, result)
    if flag_dm == 1:
        print(char_debug_print * outline_left, "MATCH", step, result, row)
    if flag_dm == 2:
        print(char_debug_print * outline_left, "STEP=BLANK, RESULT  WITHOUT NUMBER", step, result, row)
    if flag_dm == 3:
        print(" " * (number_of_spaces *
                     int(dict_routines["write_to_excel_template"][1])), "NEXT ABC RESULTS", step, result, row)
    if flag_dm == 4:
        print(" " * (number_of_spaces *
                     int(dict_routines["write_to_excel_template"][1])), "RESULT = BLANK", step, result, row)
    if flag_dm == 5:
        print(" " * (number_of_spaces *
                     int(dict_routines["write_to_excel_template"][1])), "MISSING RESULTS", step, result, row)
    if flag_dm == 6:
        print(char_debug_print * outline_left, "WHAT THE FUCK", step, result, row, "$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    if flag_dm == 7:
        print(char_debug_print * outline_left, "STEPS WITHOUT NUMBERS ???? ", step, result, row)
    print(char_routine_stop * outline_left,
          dict_routines["write_to_excel_template"][0], char_routine_stop * outline_right)


def select_regex_string(number):

    outline_left, outline_right = calculate_outline(
        dict_routines["select_regex_string"][0], dict_routines["select_regex_string"][1])
    print(char_routine_start * outline_left,
          dict_routines["select_regex_string"][0], char_routine_start * outline_right)
    print(char_debug_print * outline_left, "select_regex_string IN parameter number ", number)

    regex_string = ""
    if number == 1:
        print(char_debug_print * outline_left, "regex1")
        regex_string = re.compile(r"^1")
    if number == 2:
        print(char_debug_print * outline_left, "regex2")
        regex_string = re.compile(r"^2")
    if number == 3:
        print(char_debug_print * outline_left, "regex3")
        regex_string = re.compile(r"^3")
    if number == 4:
        print(char_debug_print * outline_left, "regex4")
        regex_string = re.compile(r"^4")
    if number == 5:
        print(char_debug_print * outline_left, "regex5")
        regex_string = re.compile(r"^5")
    if number == 6:
        print(char_debug_print * outline_left, "regex6")
        regex_string = re.compile(r"^6")
    if number == 7:
        print(char_debug_print * outline_left, "regex7")
        regex_string = re.compile(r"^7")
    if number == 8:
        print(char_debug_print * outline_left, "regex8")
        regex_string = re.compile(r"^8")
    if number == 9:
        print(char_debug_print * outline_left, "regex9")
        regex_string = re.compile(r"^9")
    if number == 10:
        print(char_debug_print * outline_left, "regex10")
        regex_string = re.compile(r"^10")
    if number == 11:
        print(char_debug_print * outline_left, "regex11")
        regex_string = re.compile(r"^11")
    if number == 12:
        print(char_debug_print * outline_left, "regex12")
        regex_string = re.compile(r"^12")
    if number == 13:
        print(char_debug_print * outline_left, "regex13")
        regex_string = re.compile(r"^13")
    if number == 14:
        print(char_debug_print * outline_left, "regex14")
        regex_string = re.compile(r"^14")
    if number == 15:
        print(char_debug_print * outline_left, "regex15")
        regex_string = re.compile(r"^15")
    if number == 16:
        print(char_debug_print * outline_left, "regex16")
        regex_string = re.compile(r"^16")
    if number == 17:
        print(char_debug_print * outline_left, "regex17")
        regex_string = re.compile(r"^17")
    if number == 18:
        print(char_debug_print * outline_left, "regex18")
        regex_string = re.compile(r"^18")
    if number == 19:
        print(char_debug_print * outline_left, "regex19")
        regex_string = re.compile(r"^19")
    if number == 20:
        print(char_debug_print * outline_left, "regex20")
        regex_string = re.compile(r"^20")
    if number == 98:
        print(char_debug_print * outline_left, "regex98")
        regex_string = re.compile(r"^98")
    if number == 99:
        print(char_debug_print * outline_left, "regex99")
        regex_string = re.compile(r"^99")

    print(char_debug_print * outline_left, "select_regex_string OUT parameter regex_string", regex_string)
    print(char_routine_stop * outline_left, dict_routines["select_regex_string"][0], char_routine_stop * outline_right)

    return regex_string


def remove_numbers(test_action):

    outline_left, outline_right = calculate_outline(
        dict_routines["remove_numbers"][0], dict_routines["remove_numbers"][1])
    print(char_routine_start * outline_left, dict_routines["remove_numbers"][0], char_routine_start * outline_right)
    print(char_debug_print * outline_left, "remove_numbers IN parameter test_action ", test_action)

    action_text = ""
    if "." in test_action:
        point_pos = test_action.find(".")
        if (point_pos > 0) and (point_pos < 4):  # Take first point after number, not ending point
            action_text = test_action[point_pos + 1:]
            print(char_debug_print * outline_left, "action_text", action_text)

    print(char_debug_print * outline_left, "remove_numbers OUT parameter action_text", action_text)
    print(char_routine_stop * outline_left, dict_routines["remove_numbers"][0], char_routine_stop * outline_right)

    return action_text


def split_steps(nr, rijnr, step, list_r, list_s, list_snumbers, list_rnumbers):

    outline_left, outline_right = calculate_outline(dict_routines["split_steps"][0], dict_routines["split_steps"][1])
    print("outline_left", outline_left)
    print("outline_right", outline_right)
    print(char_routine_start * outline_left, dict_routines["split_steps"][0], char_routine_start * outline_right)
    print(char_debug_print * outline_left, "split_steps IN parameter nr ", nr)
    print(char_debug_print * outline_left, "split_steps IN parameter rijnr ", rijnr)
    print(char_debug_print * outline_left, "split_steps IN parameter step ", step)
    print(char_debug_print * outline_left, "split_steps IN parameter list_r", list_r)
    print(char_debug_print * outline_left, "split_steps IN parameter list_snumbers", list_snumbers)
    print(char_debug_print * outline_left, "split_steps IN parameter list_rnumbers", list_rnumbers)

    write_row_start = rijnr
    write_row_stop = rijnr
    regex_sxx = select_regex_string(nr)
    list_sxx = regex_sxx.findall(step)
    print(char_debug_print * outline_left, "len(list_sxx)", len(list_sxx))
    if len(list_sxx) > 0:  # Step01
        list_chars = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l",
                      "m", "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z"]
        stepxx = step
        resultxx = ""
        regex_rxx = select_regex_string(nr)
        match_rxx = 0  # matching pair step - result
        match_rxx_char = 0  # step is present - result has second line
        match_rxx_blank = 0  # steps is present - result has blank line
        match_position = 0
        match_row = 0
        count = 0
        for y in list_r:
            print(char_debug_print * outline_left, "result = y", y)
            if (match_rxx == 1) and (len(y.strip()) == 0):
                print(char_debug_print * outline_left, "<<<skip blank y")
                break
            if "." in y:
                point_pos_result = y.find(".")
                if (point_pos_result > 0) and (point_pos_result < 4):  # Take first point after number, not ending point
                    action_text_result = "^" + y[0:point_pos_result]
                    print(char_debug_print * outline_left, "action_text_result",
                          action_text_result, "y[0:point_pos_result]", y[0:point_pos_result])
                point_pos_step = stepxx.find(".")
                if (point_pos_step > 0) and (point_pos_step < 4):  # Take first point after number, not ending point
                    action_text_step = "^" + stepxx[0:point_pos_step]
                    print(char_debug_print * outline_left, "action_text_step",
                          action_text_step, "stepxx[0:point_pos_step]", stepxx[0:point_pos_step])
                    match = re.search(action_text_step, y)
                    print(char_debug_print * outline_left, "MATCH", match, len(list_s), len(list_r))
                    print(char_debug_print * outline_left, "POSITION LEN",
                          len(stepxx[0:point_pos_step]), len(y[0:point_pos_result]))
                    if (stepxx[0:point_pos_step] != y[0:point_pos_result]) and (match is None):
                        print(char_debug_print * outline_left, "ENTERED SKIP", type(nr), nr, list_rnumbers)
                        addendum = 0
                        c = 0
                        test_string = stepxx[0:point_pos_step + 1]
                        while c < 26:
                            print(char_debug_print * outline_left, list_chars[c], test_string)
                            if list_chars[c] in test_string:
                                addendum = 1
                                c = 25
                            c += 1
                        print(char_debug_print * outline_left, "c", c, stepxx[0:point_pos_step], "addendum", addendum)
                        if (str(nr) in list_rnumbers) and (addendum == 0):
                            print(char_debug_print * outline_left, "SKIP", stepxx[0:point_pos_step], y[0:point_pos_result],
                                  len(stepxx[0:point_pos_step]), len(y[0:point_pos_result]))
                            continue
                    if (len(action_text_step) != len(action_text_result)) and (action_text_result[-1] in "0123456789") and match:
                        print(char_debug_print * outline_left, "RESET MATCH",
                              len(action_text_step), len(action_text_result))
                        match = None
                    if (point_pos_step == 3) and match:  # example 12a.
                        print(char_debug_print * outline_left,
                              "xx.a PUNT AANWEZIG  point_pos: ", point_pos_step, action_text_step, stepxx, y)
                    if (point_pos_step == 2) and match:  # example 12. or 9a.
                        print(char_debug_print * outline_left,
                              "xa. PUNT AANWEZIG  point_pos: ", point_pos_step, action_text_step, stepxx, y)
                    if (point_pos_step == 1) and match:  # example 9.
                        print(char_debug_print * outline_left,
                              "x. PUNT AANWEZIG  point_pos: ", point_pos_step, action_text_step, stepxx, y)
                    if match:
                        count += 1
                        print(char_debug_print * outline_left, "===", count, y)
                        # list_rxx = regex_rxx.findall(y)
                        # if len(list_rxx) > 0:  # Matching steps 1
                        print(char_debug_print * outline_left, "Match", y)
                        match_position = count
                        match_rxx = 1
                        resultxx = y
                        flag_debug_matches = 1
                        if action_text_step == action_text_result:
                            write_to_excel_template(stepxx, resultxx, write_row_start,
                                                    flag_remove_number, flag_debug_matches)
                        else:
                            write_row_stop += 1
                            step_blank = ""
                            write_to_excel_template(step_blank, resultxx, write_row_stop,
                                                    flag_remove_number, flag_debug_matches)
                    if match is None:  # No matching results
                        print(char_debug_print * outline_left, "NO MATCH", match)
                        print(char_debug_print * outline_left, "NO MATCH stepxx", stepxx)
                        print(char_debug_print * outline_left, "NO MATCH resultxx", resultxx)
                        regex_rxx_char = re.compile(r"^[0-9][0-9]?[a-zA-Z]")
                        list_rxx_char = regex_rxx_char.findall(y)
                        print(char_debug_print * outline_left, "NO MATCH len(list_rxx_char)", len(list_rxx_char))
                        if (len(list_rxx_char) == 0):  # RESULT = BLANK
                            resultxx = ""
                            flag_debug_matches = 4  # RESULT = BLANK
                        if y[0:point_pos_result] == "99":
                            print(char_debug_print * outline_left, "result=99", y[0:point_pos_result])
                            resultxx = y
                            flag_debug_matches = 5
                        write_to_excel_template(stepxx, resultxx, write_row_start,
                                                flag_remove_number, flag_debug_matches)
                        continue  # stop zoeken naar results
#                        if (len(list_rxx_char) == 1):  # STEP=BLANK, RESULT  WITHOUT NUMBER
#                            write_row_stop += 1
#                            resultxx = y
#                            flag_debug_matches = 2  # STEP=BLANK, RESULT  WITHOUT NUMBER
#                            write_to_excel_template(stepxx, resultxx, write_row_stop,
#                                                    flag_remove_number, flag_debug_matches)
#                            continue  # stop zoeken naar results

            if "." not in y:
                print(char_debug_print * outline_left, "PUNT NIET AANWEZIG  ")
                regex_rxx_char = re.compile(r"^[a-zA-Z]")
                list_rxx_char = regex_rxx_char.findall(y)
                if (count == match_position + 1) and (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    step_blank = ""
                    resultxx = y
                    flag_debug_matches = 2
                    write_to_excel_template(step_blank, resultxx, write_row_stop,
                                            flag_remove_number, flag_debug_matches)
                    continue  # stop zoeken naar results
                if (match_rxx_char == 1) and (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    step_blank = ""
                    resultxx = y
                    flag_debug_matches = 3
                    write_to_excel_template(step_blank, resultxx, write_row_stop,
                                            flag_remove_number, flag_debug_matches)
                    continue
                if (match_rxx_blank == 0) and (match_rxx == 0) and (len(y.strip()) == 0):  # result01 = blank
                    match_rxx_blank = 1
                    resultxx = ""
                    flag_debug_matches = 4
                    write_to_excel_template(stepxx, resultxx, write_row_start, flag_remove_number, flag_debug_matches)
                    continue
        if list_r == "":
            resultxx = ""
            flag_debug_matches = 5
            write_to_excel_template(stepxx, resultxx, write_row_start, flag_remove_number, flag_debug_matches)
        elif (match_rxx == 0):
            if nr == count:
                print(char_debug_print * outline_left, "list_r", list_r)
                print(char_debug_print * outline_left, "count", count, "nr", nr)
                resultxx = list_r[count - 1]
                flag_debug_matches = 6
                write_to_excel_template(stepxx, resultxx, write_row_start, flag_remove_number, flag_debug_matches)
    if len(list_sxx) == 0:  # Step01
        print(char_debug_print * outline_left, "NO steps")

    print(char_debug_print * outline_left,
          "split_steps OUT parameter write_row_stop", write_row_stop)
    print(char_routine_stop * outline_left, dict_routines["split_steps"][0], char_routine_stop * outline_right)

    return write_row_stop


def check_steps_and_results(read_row, read_col, row_start):

    outline_left, outline_right = calculate_outline(
        dict_routines["check_steps_and_results"][0], dict_routines["check_steps_and_results"][1])
    print(char_routine_start * outline_left,
          dict_routines["check_steps_and_results"][0], char_routine_start * outline_right)
    print(char_debug_print * outline_left, "check_steps_and_results IN parameter read_row ", read_row)
    print(char_debug_print * outline_left, "check_steps_and_results IN parameter read_col ", read_col)
    print(char_debug_print * outline_left, "check_steps_and_results IN parameter row_start ", row_start)

    row_stop = row_start
    e = tab.cell(row=read_row, column=read_col)
    if e.value is not None:

        s = tab.cell(row=read_row, column=read_col)
        r = tab.cell(row=read_row, column=read_col + 1)
        if (s.value is not None):  # Steps are present
            test_steps = s.value
            list_steps = test_steps.splitlines()
            if (r.value is not None):  # Results are present
                test_results = r.value
                list_results = test_results.splitlines()
            else:
                test_results = ""
                list_results = ""
#            if len(list_steps) != len(list_results):
            print(char_debug_print * outline_left, "S T E P S +++++++++++++>", row, len(list_steps), len(list_results))
            print(char_debug_print * outline_left, list_steps)
            print(char_debug_print * outline_left, "R E S U L T S  +++++++++>", row, len(list_steps), len(list_results))
            print(char_debug_print * outline_left, list_results)

# Steps Renumbering
            print(char_debug_print * outline_left, "------ S T E P S ----------")
            list_steps, list_stepnumbers = renumber_testcase_action(list_steps, "STEPS")
#            list_stepnumbers = list_steps[:]
            print(char_debug_print * outline_left, "R E N U M B E R  +++++++++++++>",
                  row, len(list_steps), len(list_results))
            print(char_debug_print * outline_left, "------ S T E P S ---------- list_steps", list_steps)
            print(char_debug_print * outline_left, "------ S T E P S ---------- list_stepnumbers", list_stepnumbers)
# Results Renumbering
            print(char_debug_print * outline_left, "------ R E S U L T S ----------")
            if len(list_results) != 0:
                list_results, list_resultnumbers = renumber_testcase_action(list_results, "RESULTS")
            else:
                print(char_debug_print * outline_left, "NO RESULTS------> len(list_results)", len(list_results))
                list_results = []
                list_resultnumbers = []
#            list_resultnumbers = list_results[:]
            print(char_debug_print * outline_left, "R E N U M B E R  +++++++++>", row, len(list_steps), len(list_results))
            print(char_debug_print * outline_left, "------ R E S U L T S ---------- list_results", list_results)
            print(char_debug_print * outline_left, "------ R E S U L T S ---------- list_resultnumbers", list_resultnumbers)
            cs = 0
            without_number = 0  # steps and numbers without numbers
            for x in list_steps:
                cs += 1
                print(char_debug_print * outline_left, cs, x)
                if "Precondition" in x:
                    precondition = x
                    if "." in x:
                        split_precondition_steps = x.split('.')  # Check first point
                        print(char_debug_print * outline_left, "YYYYYYYY split_precondition_steps", split_precondition_steps)
                        print(char_debug_print * outline_left,
                              "YYYYYYYY len(split_precondition_steps)", len(split_precondition_steps))
                        if (len(split_precondition_steps) > 0):
                            for i in range(0, len(split_precondition_steps)):
                                print(char_debug_print * outline_left, "i", i)
                                print(char_debug_print * outline_left,
                                      "split_precondition_steps[i] :", split_precondition_steps[i], "YYYYYYYYYYYYY")
                                if "Precondition" in split_precondition_steps[i]:
                                    print(char_debug_print * outline_left,
                                          "Precondition Catch", split_precondition_steps[i])
                                    x = split_precondition_steps[i].strip()
                                    break
                    write_col = 4  # Steps : read column F(6); Preconditions : write column E(number_of_spaces)
                    print(char_debug_print * outline_left, "Precondition ", cs, x.strip())
                    # precondition = tab.cell(row=row_start, column=write_col)
                    # precondition.value = x.strip()
                    worksheet.write(row_start, write_col, x.strip())
                    continue
                if len(x.strip()) == 0:
                    print(char_debug_print * outline_left, "Steps Blank", cs, x)
                    continue
#                    for i in range(9):
#                        stepnr = i + 1
                regex_stepnr = re.compile(r"^[0-9][0-9]?")
                list_stepnr = regex_stepnr.findall(x)
                if len(list_stepnr) == 1:
                    int_stepnr = int(list_stepnr[0])
                    print(char_debug_print * outline_left, "int_stepnr", type(int_stepnr), int_stepnr)
                    print(char_debug_print * outline_left, "split before", row_start)
                    row_start = split_steps(int_stepnr, row_start, x, list_results,
                                            list_steps, list_stepnumbers, list_resultnumbers)
                    print(char_debug_print * outline_left, "split after", row_start)
                    row_start += 1
                else:  # Steps and Results without numbers
                    if without_number == 0:
                        steps_start = row_start
                        without_number = 1
                        print(char_debug_print * outline_left, "step_start", steps_start)
                    regex_steps = re.compile(r"^[a-zA-Z][a-zA-Z]", re.MULTILINE)
                    list_stepxx = regex_steps.findall(test_steps)
                    print(char_debug_print * outline_left, "len(list_stepxx)", len(list_stepxx), list_stepxx)
                    print(char_debug_print * outline_left, "+" * 40)
                    print(char_debug_print * outline_left, "test_steps", len(test_steps), test_steps)
                    print(char_debug_print * outline_left, "+" * 40)
                    print(char_debug_print * outline_left, "test_results", test_results)
                    if (len(list_stepxx) > 0):
                        split_test_steps = test_steps.split('\n')
                        print(char_debug_print * outline_left, "XXXXXXXXX split_test_steps", split_test_steps)
                        print(char_debug_print * outline_left, "XXXXXXXXX len(split_test_steps)", len(split_test_steps))
                        c = len(list_stepxx) - 1
                        print(char_debug_print * outline_left, "c: ", c)
                        for i in range(1, len(split_test_steps)):
                            print(char_debug_print * outline_left, "i", i)
                            print(char_debug_print * outline_left,
                                  "split_test_steps[1] :", split_test_steps[1], "XXXXXXXXXXXX")
                            # print("split", split_test_steps[i])
                            print(char_debug_print * outline_left, "list_stepxx[c] :", list_stepxx[c], "XXXXXXXXXXXX")
                            if split_test_steps[i].startswith(list_stepxx[c]):
                                print(char_debug_print * outline_left, "Catch", split_test_steps[i])
                                test_steps = split_test_steps[i]
                                test_results = ""
                                break
                        stepxx = test_steps
                        resultxx = test_results
                        print(char_debug_print * outline_left, "stepxx", stepxx)
                        # if ("Precondition" not in stepxx):
                        flag_debug_matches = 7
                        write_to_excel_template(stepxx, resultxx, write_row_start,
                                                flag_remove_number, flag_debug_matches)
                        row_start += 1
        else:  # Steps are blank and results are present
            print(char_debug_print * outline_left, "STEPS ARE BLANK AND RESULTS ARE PRESENT")

    print(char_debug_print * outline_left, "check_steps_and_results OUT parameter row_star", row_start)
    print(char_routine_stop * outline_left,
          dict_routines["check_steps_and_results"][0], char_routine_stop * outline_right)

    return row_start


# ------------------MAIN -------------------------


char_routine_start = ">"
char_routine_stop = "<"
char_debug_print = "."
number_of_spaces = 2
number_char_divider = 110
flag_remove_number = 0  # Remove step or result number
flag_debug_matches = 1  # Print debug info function write_step_matches_result
dict_routines = {
    "main": ["main", 0, 4],
    "select_excel_tab": ["select_excel_tab", 1, 16],
    "write_header": ["write_header", 1, 12],
    "check_steps_and_results": ["check_steps_and_results", 1, 22],
    "renumber_testcase_action": ["renumber_testcase_action", 2, 24],
    "split_steps": ["split_steps", 2, 11],
    "write_to_excel_template": ["write_to_excel_template", 2, 23],
    "select_regex_string": ["select_regex_string", 3, 19]
}
outline_left, outline_right = calculate_outline(dict_routines["main"][0], dict_routines["main"][1])
print(char_routine_start * outline_left, dict_routines["main"][0], char_routine_start * outline_right)

list_col_access = [2, 3, 4, 5, 6]
list_col_copy = [2, 3, 4, 5]
list_col_split = [6, 7]
list_steps = []
list_header1 = ["Test Suite", "Test Cases"]
list_header2 = ["Name", "Details", "Name", "Summary", "Preconditions", "Test Execution Type",
                "Importance", "Steps", "Expected Results", "Step Execution Type", "Requirements"]
write_row_start = 2
write_row_stop = 2
list_TC_topics = ["Install & Uninstall", "First Run & Registration", "Settings",
                  "New Label and Open Label", "Printers & Consumables", "Label Editing",
                  "Import Data", "Text Based Objects", "Graphic Based Objects"]
tabnr = select_excel_tab()
inputfile = "Reviewed Test cases_DCD_Updated_19062018.xlsx"
wb = openpyxl.load_workbook(inputfile)
wb_template = xlsxwriter.Workbook('TC_template.xlsx')
worksheet = wb_template.add_worksheet("TestCases")
# wb = openpyxl.load_workbook('Test cases_DCD_FN.xlsx')
for sheet in wb.sheetnames:
    if sheet == "Totaal reviewed cases":
        continue
    if sheet == "Import Data":
        print(char_debug_print * outline_left, "SKIP TAB NO DATA----------------->", sheet, list_TC_topics[tabnr])
        continue
    if sheet == "Install & Uninstall":
        write_header()
#        continue  # discard TOC
    if sheet != list_TC_topics[tabnr]:
        print(char_debug_print * outline_left, "SKIP TAB", sheet, list_TC_topics[tabnr])
        continue
    if sheet in list_TC_topics:
        print(char_debug_print * outline_left, sheet, list_TC_topics[tabnr])
        tab = wb[sheet]
        for row in range(2, tab.max_row + 1):
            #  write_row_start += 1
            for col in range(1, tab.max_column + 1):
                if col not in list_col_access:
                    e = ""
                    continue
                e = tab.cell(row=row, column=col)
                if col in list_col_copy:
                    if e.value is not None:
                        if col == 2:  # Traceablity ID : read column B (02); Requirements : write column K(11)
                            write_col = 10
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 3:  # Functionality : read column C (03); Name : write column A(01)
                            write_col = 0
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 4:  # Task : read column D (04); Name : write column C(03)
                            write_col = 2
                            e = tab.cell(row=row, column=col)
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 5:  # Description : read column E (05); Summary : write column D(04)
                            write_col = 3
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 8:  # Priority : read column H (8); Importance : write column G(7)
                            write_col = 6
                            worksheet.write(write_row_start, write_col, e.value)
                if col in list_col_split:
                    print(char_debug_print * outline_left, "row_start", write_row_start)
                    write_row_start = check_steps_and_results(row, col, write_row_start)
wb_template.close()  # save the workbook
wb.close()
print(char_routine_stop * outline_left, dict_routines["main"][0], char_routine_stop * outline_right)
