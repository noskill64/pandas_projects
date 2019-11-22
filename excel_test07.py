import logging
import openpyxl
import re
import sys
import xlsxwriter
from openpyxl.styles import PatternFill


def calculate_outline(routine_name, routine_level):
    outline_left = number_of_spaces * int(routine_level)
    outline_right = number_char_divider - (len(routine_name) + outline_left)
    return outline_left, outline_right


def renumber_testcase_action(list_action, action_type):

    outline_left, outline_right = calculate_outline(
        dict_routines["renumber_testcase_action"][0], dict_routines["renumber_testcase_action"][1])
    # print(char_routine_start * outline_left,
    #      dict_routines["renumber_testcase_action"][0], char_routine_start * outline_right)
    str_debug = f"{char_routine_start * outline_left} {dict_routines['renumber_testcase_action'][0]} {char_routine_start * outline_right}"
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "renumber_testcase_action IN parameter list_action ", list_action)
    str_debug = f"{char_debug_print * outline_left} renumber_testcase_action IN parameter list_action {list_action}"
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "renumber_testcase_action IN parameter action_type ", action_type)
    str_debug = f"{char_debug_print * outline_left} renumber_testcase_action IN parameter action_type {action_type}"
    logger.debug(str_debug)
    letter_counter = 0
    flag_action_number = ""
    count_action_number = 100
    flag_except = 0
    no_action = 0
    c_steps = 0
    c_results = 0
    temp_list_action = []
    list_numbers = []
    if action_type == "STEPS":
        # print(char_debug_print * outline_left, "ACTION = STEPS")
        str_debug = f"{char_debug_print * outline_left} ACTION = STEPS"
        logger.debug(str_debug)

    if action_type == "RESULTS":
        # print(char_debug_print * outline_left, "ACTION = RESULTS")
        str_debug = f"{char_debug_print * outline_left} ACTION = RESULTS"
        logger.debug(str_debug)
    list_alphabet = ["a.", "b.", "c.", "d.", "f.", "g.", "h.", "i.", "j.", "k.", "l.", "m.", "n.", "o.",
                     "p.", "q.", "r.", "s.", "t.", "u.", "v.", "w.", "x.", "y.", "z."]
    for i in range(0, len(list_action)):
        if "Precondition" in list_action[i]:
            continue
        # print(char_debug_print * outline_left, "Action", "i", i, "len(list_action)", len(list_action), list_action[i])
        str_debug = f"{char_debug_print * outline_left} Action i {i} len(list_action) {len(list_action)} {list_action[i]}"
        logger.debug(str_debug)
        point_pos = list_action[i].find(".")
        # print(char_debug_print * outline_left, "point_pos", point_pos)
        str_debug = f"{char_debug_print * outline_left} point_pos {point_pos}"
        logger.debug(str_debug)
        if (point_pos == -1) and (len(list_action[i].strip()) == 0):  # skip blank lines
            # print(char_debug_print * outline_left,
            #      "skip blank line <---------------------------", "len(list_action[i])", len(list_action[i].strip()))
            str_debug = f"{char_debug_print * outline_left} skip blank line \
                < --------------------------- len(list_action[i]) {len(list_action[i].strip())}"
            logger.debug(str_debug)
            continue
        # Take first point after number, not ending point)
        if ("." in list_action[i]) and ((point_pos > 0) and (point_pos < 4)):
            split_number_action = list_action[i].split('.')  # Check first point
            action_number = split_number_action[0]
            temp_number = split_number_action[0]
            letter_counter = 0
            no_action = 1
            flag_action_number = action_number
            # print(char_debug_print * outline_left, "Action number", action_number, split_number_action)
            str_debug = f"{char_debug_print * outline_left} Action number \
                {action_number} {split_number_action}"
            logger.debug(str_debug)
        else:
            # print(char_debug_print * outline_left, "blabla", list_action[i], "no_action", no_action,
                #  "len(list_action[i].strip()", len(list_action[i].strip()))
            str_debug = f"{char_debug_print * outline_left} blabla \
                {list_action[i]} no action {no_action} len(list_action[i].strip() {len(list_action[i].strip())}"
            logger.debug(str_debug)
            if (no_action == 0) and (len(list_action[i].strip()) == 0):  # skip first blank step line
                # print(char_debug_print * outline_left, "skip first blank step line")
                str_debug = f"{char_debug_print * outline_left} skip first blank step line"
                logger.debug(str_debug)
                continue
            if len(list_action[i].strip()) == 0:  # Blanks
                if flag_action_number != action_number:
                    flag_action_number = action_number
                    letter_counter = 0
                    action_number_next, temp_number = action_number + list_alphabet[letter_counter]
                else:
                    action_number_next, temp_number = action_number + list_alphabet[letter_counter]
                    letter_counter += 1
                    # list_steps[i] = step_number_next + list_steps[i]
                    # print(char_debug_print * outline_left, "Steps Blank", list_action[i])
                    str_debug = f"{char_debug_print * outline_left} Steps Blank {list_action[i]}"
                    logger.debug(str_debug)
            else:  # Starts with letter, no step number
                regex_action = re.compile(r"^[a-zA-Z][a-zA-Z]", re.MULTILINE)
                list_char_detect = regex_action.findall(list_action[i])
                if (len(list_char_detect) > 0):
                    try:
                        dummy = len(action_number)
                    except UnboundLocalError:
                        # print(char_debug_print * outline_left, "EXCEPT")
                        str_debug = f"{char_debug_print * outline_left} EXCEPT"
                        logger.debug(str_debug)
                        flag_except = 1
                        count_action_number -= 1
                        action_number, temp_number = str(count_action_number)
                        # print(char_debug_print * outline_left, "action_number", action_number)
                        str_debug = f"{char_debug_print * outline_left} action_number {action_number}"
                        logger.debug(str_debug)
                        pass
                    if flag_action_number != action_number:
                        # print(char_debug_print * outline_left, "THEN")
                        str_debug = f"{char_debug_print * outline_left} THEN"
                        logger.debug(str_debug)
                        flag_action_number = action_number
                        letter_counter = 0
                        action_number_next = action_number + list_alphabet[letter_counter]
                        temp_number = action_number + list_alphabet[letter_counter]
                        if (action_number == "99") and (flag_except == 1):
                            # print(char_debug_print * outline_left,
                            #      "action_number == 99", action_number, "flag_except", flag_except)
                            str_debug = f"{char_debug_print * outline_left} action_number == 99 {action_number} flag_except {flag_except} "
                            logger.debug(str_debug)
                            action_number_next, temp_number = "99. "
                            flag_except = 0
                        else:
                            count_action_number -= 1
                            action_number = str(count_action_number)
                            # print(char_debug_print * outline_left,
                            #      "action_number == XX", action_number, "flag_except", flag_except)
                            str_debug = f"{char_debug_print * outline_left} action_number == XX {action_number} flag_except {flag_except}"
                            logger.debug(str_debug)
                            action_number_next = action_number + ". "
                            temp_number = action_number + ". "
                            flag_except = 0
                        # print(char_debug_print * outline_left, "action_number_next", action_number_next, i)
                        str_debug = f"{char_debug_print * outline_left} action_number_next {action_number_next} i {i}"
                        logger.debug(str_debug)
                    else:
                        # print(char_debug_print * outline_left, "ELSE")
                        str_debug = f"{char_debug_print * outline_left} ELSE"
                        logger.debug(str_debug)
                        action_number_next = action_number + list_alphabet[letter_counter]
                        temp_number = action_number + list_alphabet[letter_counter]
                        letter_counter += 1
                    list_action[i] = action_number_next + list_action[i]
                    # print(char_debug_print * outline_left, "Action char", list_action[i])
                    str_debug = f"{char_debug_print * outline_left} Action char {list_action[i]}"
                    logger.debug(str_debug)
                    action_number_next = ""
        # print(char_debug_print * outline_left, "i", i, "list_action[i]", list_action[i])
        str_debug = f"{char_debug_print * outline_left} i {i} list_action[i] {list_action[i]}"
        logger.debug(str_debug)
        temp_list_action.append(list_action[i])
        if action_type == "STEPS":
            # print(char_debug_print * outline_left, "c_steps", c_steps, "c_results", c_results)
            str_debug = f"{char_debug_print * outline_left} c_steps {c_steps} c_results {c_results}"
            logger.debug(str_debug)
            list_numbers.append(temp_number)
            c_steps += 1
        if action_type == "RESULTS":
            # print(char_debug_print * outline_left, "ACTION = RESULTS")
            str_debug = f"{char_debug_print * outline_left} ACTION = RESULTS"
            logger.debug(str_debug)
            list_numbers.append(temp_number)
            c_results += 1
    list_action.clear()
    list_action = list(temp_list_action)
    if action_type == "STEPS":
        # print(char_debug_print * outline_left, "XX step numbers XX")
        str_debug = f"{char_debug_print * outline_left} XX step numbers XX"
        logger.debug(str_debug)
        # print(char_debug_print * outline_left, list_numbers)
        str_debug = f"{char_debug_print * outline_left} list_numbers {list_numbers}"
        logger.debug(str_debug)
    if action_type == "RESULTS":
        # print(char_debug_print * outline_left, "XX result numbers XX")
        str_debug = f"{char_debug_print * outline_left} XX result numbers XX"
        logger.debug(str_debug)
        # print(char_debug_print * outline_left, list_numbers)
        str_debug = f"{char_debug_print * outline_left} list_numbers {list_numbers}"
        logger.debug(str_debug)
    # print(char_debug_print * outline_left, "renumber_testcase_action OUT parameter list_action ", list_action)
    str_debug = f"{char_debug_print * outline_left} renumber_testcase_action OUT parameter list_action {list_action}"
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "renumber_testcase_action OUT parameter list_numbers ", list_numbers)
    str_debug = f"{char_debug_print * outline_left} renumber_testcase_action OUT parameter list_numbers {list_numbers}"
    logger.debug(str_debug)
    # print(char_routine_stop * outline_left,
    #      dict_routines["renumber_testcase_action"][0], char_routine_stop * outline_right)
    str_debug = f"{char_routine_stop * outline_left} {dict_routines['renumber_testcase_action'][0]} {char_routine_stop * outline_right} "
    logger.debug(str_debug)
    return list_action, list_numbers


def select_excel_tab():

    outline_left, outline_right = calculate_outline(
        dict_routines["select_excel_tab"][0], dict_routines["select_excel_tab"][1])
    # print(char_routine_start * outline_left, dict_routines["select_excel_tab"][0], char_routine_start * outline_right)
    str_debug = f"{char_routine_start * outline_left} {dict_routines['select_excel_tab'][0]} {char_routine_start * outline_right} "
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "Number of arguments: ", len(sys.argv), "arguments")
    str_debug = f"{char_debug_print * outline_left} Number of arguments {len(sys.argv)} arguments"
    logger.debug(str_debug)
    if len(sys.argv) > 1:
        str_D001 = char_debug_print * outline_left, "D001 Argument list: ", str(sys.argv), type(sys.argv)
        # print(char_debug_print * outline_left, str_D001)
        str_debug = f"{char_debug_print * outline_left} {str_D001}"
        logger.debug(str_debug)
        # print(char_debug_print * outline_left, "D001 Argument list: ", str(sys.argv), type(sys.argv))
        str_debug = f"{char_debug_print * outline_left} D001 Argument list: {str(sys.argv)} {type(sys.argv)}"
        logger.debug(str_debug)
        if len(sys.argv) == 2:
            if sys.argv[1].isnumeric():
                # print(char_debug_print * outline_left, "D002 isnumeric", sys.argv[1])
                str_debug = f"{char_debug_print * outline_left} D002 isnumeric {sys.argv[1]}"
                logger.debug(str_debug)
                number_arg = int(sys.argv[1])
                if (number_arg > 0) and (number_arg <= 9):
                    number_arg -= 1
                    # print(char_debug_print * outline_left, "D003", number_arg, list_TC_topics[number_arg])
                    str_debug = f"{char_debug_print * outline_left} D003 {number_arg} {number_arg, list_TC_topics[number_arg]}"
                    logger.debug(str_debug)
        else:
            for number_arg in range(1, len(sys.argv)):
                # print(char_debug_print * outline_left, "D004", number_arg, list_TC_topics[number_arg])
                str_debug = f"{char_debug_print * outline_left} D004 {number_arg} {number_arg, list_TC_topics[number_arg]}"
                logger.debug(str_debug)
    # print(char_debug_print * outline_left, "select_excel_tab OUT parameter number_arg ", number_arg)
    str_debug = f"{char_debug_print * outline_left} select_excel_tab OUT parameter number_arg {number_arg}"
    logger.debug(str_debug)
    # print(char_routine_stop * outline_left, dict_routines["select_excel_tab"][0], char_routine_stop * outline_right)
    str_debug = f"{char_routine_stop * outline_left} {dict_routines['select_excel_tab'][0]} {char_routine_stop * outline_right}"
    logger.debug(str_debug)
    return number_arg


def write_header():

    outline_left, outline_right = calculate_outline(
        dict_routines["write_header"][0], dict_routines["write_header"][1])
    # print(char_routine_start * outline_left, dict_routines["write_header"][0], char_routine_start * outline_right)
    str_debug = f"{char_routine_start * outline_left} {dict_routines['write_header'][0]} {char_routine_start * outline_right} "
    logger.debug(str_debug)
    #  Print header 1
    header1_row = 0
    header1_col1 = 0
    header1_col2 = 2
    # Create a format to use in the merged range.
    merge_format = wb_template.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '808080'})
    # Merge 2 cells.
    # print(char_debug_print * outline_left, "D005", "Header1", header1_row,
    #      header1_col1, list_header1[0])  # Header 1 Column Test Suite
    str_debug = f"{char_debug_print * outline_left} D005  Header1{header1_row} {header1_col1} {list_header1[0]}"
    logger.debug(str_debug)
    worksheet.write(header1_row, header1_col1, list_header1[0])
    worksheet.merge_range('A1:B1', list_header1[0], merge_format)
    # Merge 9 cells.
    # print(char_debug_print * outline_left, "D006", "Header1", header1_row,
    #      header1_col2, list_header1[1])  # Header 1 Column Test Cases
    str_debug = f"{char_debug_print * outline_left} D006 Header1{header1_row} {header1_col2} {list_header1[1]}"
    logger.debug(str_debug)
    worksheet.write(header1_row, header1_col2, list_header1[1])
    worksheet.merge_range('C1:K1', list_header1[1], merge_format)  # format gray text
    #  print header 2
    header2_row = 1
    header2_max_col = 11
    for header2_col in range(header2_max_col):
        # print(char_debug_print * outline_left, "D007", "Header2", header2_row, header2_col, list_header2[header2_col])
        str_debug = f"{char_debug_print * outline_left} D007 Header2 {header2_row} {header2_col} {list_header2[header2_col]}"
        logger.debug(str_debug)
        #    worksheet.cell(header2_row, header2_col).fill = PatternFill(fgColor='808080', fill_type='solid')
        worksheet.write(header2_row, header2_col, list_header2[header2_col], merge_format)
        #    my_fill = PatternFill(fgColor='FFFFFF', bgColor='FFFFFF', fill_type='solid')
        #    worksheet['A2'].fill = my_fill
    # print(char_routine_stop * outline_left, dict_routines["write_header"][0], char_routine_stop * outline_right)
    str_debug = f"{char_routine_stop * outline_left} {dict_routines['write_header'][0]} {char_routine_stop * outline_right}"
    logger.debug(str_debug)


def write_to_excel_template(step, result, row, flag_rn, flag_dm):

    outline_left, outline_right = calculate_outline(
        dict_routines["write_to_excel_template"][0], dict_routines["write_to_excel_template"][1])
    # print(char_routine_start * outline_left,
    #      dict_routines["write_to_excel_template"][0], char_routine_start * outline_right)
    str_debug = f"{char_routine_start * outline_left} {dict_routines['write_to_excel_template'][0]} {char_routine_start * outline_right} "
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "write_to_excel_template IN parameter step ", step)
    str_debug = f"{char_debug_print * outline_left} write_to_excel_template IN parameter step{step}"
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "write_to_excel_template IN parameter result ", result)
    str_debug = f"{char_debug_print * outline_left} write_to_excel_template IN parameter result {result}"
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "write_to_excel_template IN parameter row ", row)
    str_debug = f"{char_debug_print * outline_left} write_to_excel_template IN parameter row {row}"
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "write_to_excel_template IN parameter flag_rn ", flag_rn)
    str_debug = f"{char_debug_print * outline_left} write_to_excel_template IN parameter flag_rn {flag_rn}"
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "write_to_excel_template IN parameter flag_dm ", flag_dm)
    str_debug = f"{char_debug_print * outline_left} write_to_excel_template IN parameter flag_dm {flag_dm}"
    logger.debug(str_debug)

    col = 7  # Steps : read column F(6); Steps : write column H(8)
    if flag_remove_number == 1:
        step = remove_numbers(step)
    worksheet.write(row, col, step)
    col = 8  # Steps : read column F(6); Steps : write column I(9)
    if flag_remove_number == 1:
        result = remove_numbers(result)
    worksheet.write(row, col, result)
    if flag_dm == 1:
        # print(char_debug_print * outline_left, "MATCH", step, result, row)
        str_debug = f"{char_debug_print * outline_left} MATCH {step} {result}{row} "
        logger.debug(str_debug)
    if flag_dm == 2:
        # print(char_debug_print * outline_left, "STEP=BLANK, RESULT  WITHOUT NUMBER", step, result, row)
        str_debug = f"{char_debug_print * outline_left} STEP=BLANK, RESULT  WITHOUT NUMBER {step} {result}{row} "
        logger.debug(str_debug)
    if flag_dm == 3:
        # print(" " * (number_of_spaces *
        #             int(dict_routines["write_to_excel_template"][1])), "NEXT ABC RESULTS", step, result, row)
        str_debug = f"{' ' * (number_of_spaces * int(dict_routines['write_to_excel_template'][1]))} NEXT ABC RESULTS {step} {result} {row} "
        logger.debug(str_debug)
    if flag_dm == 4:
        # print(" " * (number_of_spaces *
        #             int(dict_routines["write_to_excel_template"][1])), "RESULT = BLANK", step, result, row)
        str_debug = f"{' ' * (number_of_spaces * int(dict_routines['write_to_excel_template'][1]))} RESULT = BLANK {step} {result} {row} "
        logger.debug(str_debug)
    if flag_dm == 5:
        # print(" " * (number_of_spaces *
        #             int(dict_routines["write_to_excel_template"][1])), "MISSING RESULTS", step, result, row)
        str_debug = f"{' ' * (number_of_spaces * int(dict_routines['write_to_excel_template'][1]))} MISSING RESULTS {step} {result} {row} "
        logger.debug(str_debug)
    if flag_dm == 6:
        # print(char_debug_print * outline_left, "WHAT THE FUCK", step, result, row, "$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
        str_debug = f"{char_debug_print * outline_left} WHAT THE FUCK {step} {result}{row} "
        logger.debug(str_debug)
    if flag_dm == 7:
        # print(char_debug_print * outline_left, "STEPS WITHOUT NUMBERS ???? ", step, result, row)
        str_debug = f"{char_debug_print * outline_left} STEPS WITHOUT NUMBERS ???? {step} {result}{row} "
        logger.debug(str_debug)
    # print(char_routine_stop * outline_left,
    #      dict_routines["write_to_excel_template"][0], char_routine_stop * outline_right)
    str_debug = f"{char_routine_stop * outline_left} {dict_routines['write_to_excel_template'][0]} {char_routine_stop * outline_right}"
    logger.debug(str_debug)


def remove_numbers(test_action):

    outline_left, outline_right = calculate_outline(
        dict_routines["remove_numbers"][0], dict_routines["remove_numbers"][1])
    # print(char_routine_start * outline_left, dict_routines["remove_numbers"][0], char_routine_start * outline_right)
    str_debug = f"{char_routine_start * outline_left} {dict_routines['remove_numbers'][0]} {char_routine_start * outline_right} "
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "remove_numbers IN parameter test_action ", test_action)
    str_debug = f"{char_debug_print * outline_left} remove_numbers IN parameter test_action  {test_action} "
    logger.debug(str_debug)

    action_text = ""
    if "." in test_action:
        point_pos = test_action.find(".")
        if (point_pos > 0) and (point_pos < 4):  # Take first point after number, not ending point
            action_text = test_action[point_pos + 1:]
            # print(char_debug_print * outline_left, "action_text", action_text)
            str_debug = f"{char_debug_print * outline_left} action_text {action_text} "
            logger.debug(str_debug)

    # print(char_debug_print * outline_left, "remove_numbers OUT parameter action_text", action_text)
    str_debug = f"{char_debug_print * outline_left} remove_numbers OUT parameter action_text {action_text} "
    logger.debug(str_debug)
    # print(char_routine_stop * outline_left, dict_routines["remove_numbers"][0], char_routine_stop * outline_right)
    str_debug = f"{char_routine_stop * outline_left} {dict_routines['remove_numbers'][0]} {char_routine_stop * outline_right}"
    logger.debug(str_debug)

    return action_text


def split_steps(nr, rijnr, step, list_r, list_s, list_snumbers, list_rnumbers):

    outline_left, outline_right = calculate_outline(dict_routines["split_steps"][0], dict_routines["split_steps"][1])
    # print("outline_left", outline_left)
    # print("outline_right", outline_right)
    # print(char_routine_start * outline_left, dict_routines["split_steps"][0], char_routine_start * outline_right)
    str_debug = f"{char_routine_start * outline_left} {dict_routines['split_steps'][0]} {char_routine_start * outline_right} "
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "split_steps IN parameter nr ", nr)
    str_debug = f"{char_debug_print * outline_left} split_steps IN parameter nr  {nr} "
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "split_steps IN parameter rijnr ", rijnr)
    str_debug = f"{char_debug_print * outline_left} split_steps IN parameter rijnr  {rijnr} "
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "split_steps IN parameter step ", step)
    str_debug = f"{char_debug_print * outline_left} split_steps IN parameter step  {step} "
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "split_steps IN parameter list_r", list_r)
    str_debug = f"{char_debug_print * outline_left} split_steps IN parameter list_r  {list_r} "
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "split_steps IN parameter list_snumbers", list_snumbers)
    str_debug = f"{char_debug_print * outline_left} split_steps IN parameter list_snumbers  {list_snumbers} "
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "split_steps IN parameter list_rnumbers", list_rnumbers)
    str_debug = f"{char_debug_print * outline_left} split_steps IN parameter list_rnumbers  {list_rnumbers} "
    logger.debug(str_debug)

    write_row_start = rijnr
    write_row_stop = rijnr
    regex_sxx = re.compile(r"^" + str(nr))
    # print(char_debug_print * outline_left, "regex_sxx", regex_sxx)
    str_debug = f"{char_debug_print * outline_left} regex_sxx  {regex_sxx} "
    logger.debug(str_debug)
    list_sxx = regex_sxx.findall(step)
    # print(char_debug_print * outline_left, "len(list_sxx)", len(list_sxx))
    str_debug = f"{char_debug_print * outline_left} len(list_sxx)  {len(list_sxx)} "
    logger.debug(str_debug)
    if len(list_sxx) > 0:  # Step01
        list_chars = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l",
                      "m", "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z"]
        stepxx = step
        resultxx = ""
        regex_rxx = re.compile(r"^" + str(nr))
        match_rxx = 0  # matching pair step - result
        match_rxx_char = 0  # step is present - result has second line
        match_rxx_blank = 0  # steps is present - result has blank line
        match_position = 0
        match_row = 0
        count = 0
        for y in list_r:
            # print(char_debug_print * outline_left, "result = y", y)
            str_debug = f"{char_debug_print * outline_left} result = y  {y} "
            logger.debug(str_debug)
            if (match_rxx == 1) and (len(y.strip()) == 0):
                # print(char_debug_print * outline_left, "<<<skip blank y")
                str_debug = f"{char_debug_print * outline_left} <<<skip blank y "
                logger.debug(str_debug)
                break
            if "." in y:
                point_pos_result = y.find(".")
                if (point_pos_result > 0) and (point_pos_result < 4):  # Take first point after number, not ending point
                    action_text_result = "^" + y[0:point_pos_result]
                    # print(char_debug_print * outline_left, "action_text_result",
                    #      action_text_result, "y[0:point_pos_result]", y[0:point_pos_result])
                    str_debug = f"{char_debug_print * outline_left} action_text_result \
                                        {action_text_result} y[0:point_pos_result] {y[0:point_pos_result]}"
                    logger.debug(str_debug)
                point_pos_step = stepxx.find(".")
                if (point_pos_step > 0) and (point_pos_step < 4):  # Take first point after number, not ending point
                    action_text_step = "^" + stepxx[0:point_pos_step]
                    # print(char_debug_print * outline_left, "action_text_step",
                    #      action_text_step, "stepxx[0:point_pos_step]", stepxx[0:point_pos_step])
                    str_debug = f"{char_debug_print * outline_left} action_text_step  \
                                        {action_text_step} stepxx[0:point_pos_step] {stepxx[0:point_pos_step]}"
                    logger.debug(str_debug)
                    match = re.search(action_text_step, y)
                    # print(char_debug_print * outline_left, "MATCH", match, len(list_s), len(list_r))
                    str_debug = f"{char_debug_print * outline_left} MATCH {match} {len(list_s)} {len(list_r)}"
                    logger.debug(str_debug)
                    # print(char_debug_print * outline_left, "POSITION LEN",
                    #      len(stepxx[0:point_pos_step]), len(y[0:point_pos_result]))
                    str_debug = f"{char_debug_print * outline_left} POSITION LEN  \
                                        {len(stepxx[0:point_pos_step])} {len(y[0:point_pos_result])}"
                    logger.debug(str_debug)
                    if (stepxx[0:point_pos_step] != y[0:point_pos_result]) and (match is None):
                        # print(char_debug_print * outline_left, "ENTERED SKIP", type(nr), nr, list_rnumbers)
                        str_debug = f"{char_debug_print * outline_left} ENTERED SKIP  \
                                        {type(nr)} {nr} {list_rnumbers}"
                        logger.debug(str_debug)
                        addendum = 0
                        c = 0
                        test_string = stepxx[0:point_pos_step + 1]
                        while c < 26:
                            # print(char_debug_print * outline_left, list_chars[c], test_string)
                            str_debug = f"{char_debug_print * outline_left} {list_chars[c]} {test_string}"
                            logger.debug(str_debug)

                            if list_chars[c] in test_string:
                                addendum = 1
                                c = 25
                            c += 1
                        # print(char_debug_print * outline_left, "c", c, stepxx[0:point_pos_step], "addendum", addendum)
                        str_debug = f"{char_debug_print * outline_left} c  {c} {stepxx[0:point_pos_step]}\
                                        addendum {addendum}"
                        logger.debug(str_debug)
                        if (str(nr) in list_rnumbers) and (addendum == 0):
                            # print(char_debug_print * outline_left, "SKIP", stepxx[0:point_pos_step], y[0:point_pos_result],
                            #      len(stepxx[0:point_pos_step]), len(y[0:point_pos_result]))
                            str_debug = f"{char_debug_print * outline_left} SKIP  {stepxx[0:point_pos_step]} {y[0:point_pos_result]}\
                                        {len(stepxx[0:point_pos_step])} {len(y[0:point_pos_result])}"
                            logger.debug(str_debug)
                            continue
                    if (len(action_text_step) != len(action_text_result)) and (action_text_result[-1] in "0123456789") and match:
                        # print(char_debug_print * outline_left, "RESET MATCH",
                        #      len(action_text_step), len(action_text_result))
                        str_debug = f"{char_debug_print * outline_left} RESET MATCH  {len(action_text_step)} {len(action_text_result)}"
                        logger.debug(str_debug)
                        match = None
                    if (point_pos_step == 3) and match:  # example 12a.
                        # print(char_debug_print * outline_left,
                        #      "xx.a PUNT AANWEZIG  point_pos: ", point_pos_step, action_text_step, stepxx, y)
                        str_debug = f"{char_debug_print * outline_left} xx.a PUNT AANWEZIG  point_pos:  {point_pos_step} {action_text_step}\
                                    {stepxx} {y}"
                        logger.debug(str_debug)
                    if (point_pos_step == 2) and match:  # example 12. or 9a.
                        # print(char_debug_print * outline_left,
                        #      "xa. PUNT AANWEZIG  point_pos: ", point_pos_step, action_text_step, stepxx, y)
                        str_debug = f"{char_debug_print * outline_left} xa. PUNT AANWEZIG  point_pos: {point_pos_step} {action_text_step}\
                                    {stepxx} {y}"
                        logger.debug(str_debug)
                    if (point_pos_step == 1) and match:  # example 9.
                        # print(char_debug_print * outline_left,
                        #      "x. PUNT AANWEZIG  point_pos: ", point_pos_step, action_text_step, stepxx, y)
                        str_debug = f"{char_debug_print * outline_left} x. PUNT AANWEZIG  point_pos: {point_pos_step} {action_text_step}\
                                    {stepxx} {y}"
                        logger.debug(str_debug)
                    if match:
                        count += 1
                        # print(char_debug_print * outline_left, "===", count, y)
                        str_debug = f"{char_debug_print * outline_left} ===  {count} {y}"
                        logger.debug(str_debug)
                        # list_rxx = regex_rxx.findall(y)
                        # if len(list_rxx) > 0:  # Matching steps 1
                        # print(char_debug_print * outline_left, "Match", y)
                        str_debug = f"{char_debug_print * outline_left} Match  {y}"
                        logger.debug(str_debug)
                        match_position = count
                        match_rxx = 1
                        resultxx = y
                        flag_debug_matches = 1
                        if action_text_step == action_text_result:
                            write_to_excel_template(stepxx, resultxx, write_row_start,
                                                    flag_remove_number, flag_debug_matches)
                        else:
                            write_row_stop += 1
                            step_blank = ""
                            write_to_excel_template(step_blank, resultxx, write_row_stop,
                                                    flag_remove_number, flag_debug_matches)
                    if match is None:  # No matching results
                        # print(char_debug_print * outline_left, "NO MATCH", match)
                        str_debug = f"{char_debug_print * outline_left} NO MATCH {match}"
                        logger.debug(str_debug)
                        # print(char_debug_print * outline_left, "NO MATCH stepxx", stepxx)
                        str_debug = f"{char_debug_print * outline_left} NO MATCH stepxx {stepxx}"
                        logger.debug(str_debug)
                        # print(char_debug_print * outline_left, "NO MATCH resultxx", resultxx)
                        str_debug = f"{char_debug_print * outline_left} NO MATCH resultxx {resultxx}"
                        logger.debug(str_debug)
                        regex_rxx_char = re.compile(r"^[0-9][0-9]?[a-zA-Z]")
                        list_rxx_char = regex_rxx_char.findall(y)
                        # print(char_debug_print * outline_left, "NO MATCH len(list_rxx_char)", len(list_rxx_char))
                        str_debug = f"{char_debug_print * outline_left} NO MATCH len(list_rxx_char) {len(list_rxx_char)}"
                        logger.debug(str_debug)
                        if (len(list_rxx_char) == 0):  # RESULT = BLANK
                            resultxx = ""
                            flag_debug_matches = 4  # RESULT = BLANK
                        if y[0:point_pos_result] == "99":
                            # print(char_debug_print * outline_left, "result=99", y[0:point_pos_result])
                            str_debug = f"{char_debug_print * outline_left} result=99 {y[0:point_pos_result]}"
                            logger.debug(str_debug)
                            resultxx = y
                            flag_debug_matches = 5
                        write_to_excel_template(stepxx, resultxx, write_row_start,
                                                flag_remove_number, flag_debug_matches)
                        continue  # stop zoeken naar results
#                        if (len(list_rxx_char) == 1):  # STEP=BLANK, RESULT  WITHOUT NUMBER
#                            write_row_stop += 1
#                            resultxx = y
#                            flag_debug_matches = 2  # STEP=BLANK, RESULT  WITHOUT NUMBER
#                            write_to_excel_template(stepxx, resultxx, write_row_stop,
#                                                    flag_remove_number, flag_debug_matches)
#                            continue  # stop zoeken naar results

            if "." not in y:
                # print(char_debug_print * outline_left, "PUNT NIET AANWEZIG  ")
                str_debug = f"{char_debug_print * outline_left} PUNT NIET AANWEZIG "
                logger.debug(str_debug)
                regex_rxx_char = re.compile(r"^[a-zA-Z]")
                list_rxx_char = regex_rxx_char.findall(y)
                if (count == match_position + 1) and (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    step_blank = ""
                    resultxx = y
                    flag_debug_matches = 2
                    write_to_excel_template(step_blank, resultxx, write_row_stop,
                                            flag_remove_number, flag_debug_matches)
                    continue  # stop zoeken naar results
                if (match_rxx_char == 1) and (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    step_blank = ""
                    resultxx = y
                    flag_debug_matches = 3
                    write_to_excel_template(step_blank, resultxx, write_row_stop,
                                            flag_remove_number, flag_debug_matches)
                    continue
                if (match_rxx_blank == 0) and (match_rxx == 0) and (len(y.strip()) == 0):  # result01 = blank
                    match_rxx_blank = 1
                    resultxx = ""
                    flag_debug_matches = 4
                    write_to_excel_template(stepxx, resultxx, write_row_start, flag_remove_number, flag_debug_matches)
                    continue
        if list_r == "":
            resultxx = ""
            flag_debug_matches = 5
            write_to_excel_template(stepxx, resultxx, write_row_start, flag_remove_number, flag_debug_matches)
        elif (match_rxx == 0):
            if nr == count:
                # print(char_debug_print * outline_left, "list_r", list_r)
                str_debug = f"{char_debug_print * outline_left} list_r {list_r} "
                logger.debug(str_debug)
                # print(char_debug_print * outline_left, "count", count, "nr", nr)
                str_debug = f"{char_debug_print * outline_left} count {count} nr {nr} "
                logger.debug(str_debug)
                resultxx = list_r[count - 1]
                flag_debug_matches = 6
                write_to_excel_template(stepxx, resultxx, write_row_start, flag_remove_number, flag_debug_matches)
    if len(list_sxx) == 0:  # Step01
        # print(char_debug_print * outline_left, "NO steps")
        str_debug = f"{char_debug_print * outline_left} NO steps"
        logger.debug(str_debug)

    # print(char_debug_print * outline_left,
    #      "split_steps OUT parameter write_row_stop", write_row_stop)
    str_debug = f"{char_debug_print * outline_left} split_steps OUT parameter write_row_stop {write_row_stop}"
    logger.debug(str_debug)
    # print(char_routine_stop * outline_left, dict_routines["split_steps"][0], char_routine_stop * outline_right)
    str_debug = f"{char_routine_stop * outline_left} {dict_routines['split_steps'][0]} {char_routine_stop * outline_right} "
    logger.debug(str_debug)

    return write_row_stop


def check_steps_and_results(read_row, read_col, row_start):

    outline_left, outline_right = calculate_outline(
        dict_routines["check_steps_and_results"][0], dict_routines["check_steps_and_results"][1])
    # print(char_routine_start * outline_left,
    #      dict_routines["check_steps_and_results"][0], char_routine_start * outline_right)
    str_debug = f"{char_routine_start * outline_left} {dict_routines['check_steps_and_results'][0]} {char_routine_start * outline_right} "
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "check_steps_and_results IN parameter read_row ", read_row)
    str_debug = f"{char_debug_print * outline_left} check_steps_and_results IN parameter read_row  {read_row}"
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "check_steps_and_results IN parameter read_col ", read_col)
    str_debug = f"{char_debug_print * outline_left} check_steps_and_results IN parameter read_col {read_col}"
    logger.debug(str_debug)
    # print(char_debug_print * outline_left, "check_steps_and_results IN parameter row_start ", row_start)
    str_debug = f"{char_debug_print * outline_left} check_steps_and_results IN parameter row_start {row_start}"
    logger.debug(str_debug)

    row_stop = row_start
    e = tab.cell(row=read_row, column=read_col)
    if e.value is not None:

        s = tab.cell(row=read_row, column=read_col)
        r = tab.cell(row=read_row, column=read_col + 1)
        if (s.value is not None):  # Steps are present
            test_steps = s.value
            list_steps = test_steps.splitlines()
            if (r.value is not None):  # Results are present
                test_results = r.value
                list_results = test_results.splitlines()
            else:
                test_results = ""
                list_results = ""
#            if len(list_steps) != len(list_results):
            # print(char_debug_print * outline_left, "S T E P S +++++++++++++>", row, len(list_steps), len(list_results))
            str_debug = f"{char_debug_print * outline_left} S T E P S +++++++++++++> {row} \
                            {len(list_steps)} {len(list_results)}"
            logger.debug(str_debug)
            # print(char_debug_print * outline_left, list_steps)
            str_debug = f"{char_debug_print * outline_left} {list_steps}"
            logger.debug(str_debug)
            # print(char_debug_print * outline_left, "R E S U L T S  +++++++++>", row, len(list_steps), len(list_results))
            str_debug = f"{char_debug_print * outline_left} R E S U L T S  +++++++++> {row} \
                            {len(list_steps)} {len(list_results)}"
            logger.debug(str_debug)
            # print(char_debug_print * outline_left, list_results)
            str_debug = f"{char_debug_print * outline_left} {list_results}"
            logger.debug(str_debug)

# Steps Renumbering
            # print(char_debug_print * outline_left, "------ S T E P S ----------")
            str_debug = f"{char_debug_print * outline_left} ------ S T E P S ----------"
            logger.debug(str_debug)
            list_steps, list_stepnumbers = renumber_testcase_action(list_steps, "STEPS")
#            list_stepnumbers = list_steps[:]
            # print(char_debug_print * outline_left, "R E N U M B E R  +++++++++++++>",
            #      row, len(list_steps), len(list_results))
            str_debug = f"{char_debug_print * outline_left} R E S U L T S  +++++++++> {row} \
                            {len(list_steps)} {len(list_results)}"
            # print(char_debug_print * outline_left, "------ S T E P S ---------- list_steps", list_steps)
            str_debug = f"{char_debug_print * outline_left} ------ S T E P S ---------- list_steps {list_steps}"
            logger.debug(str_debug)
            # print(char_debug_print * outline_left, "------ S T E P S ---------- list_stepnumbers", list_stepnumbers)
            str_debug = f"{char_debug_print * outline_left} ------ S T E P S ---------- list_stepnumbers {list_stepnumbers}"
            logger.debug(str_debug)
# Results Renumbering
            # print(char_debug_print * outline_left, "------ R E S U L T S ----------")
            str_debug = f"{char_debug_print * outline_left} ------ R E S U L T S ---------"
            logger.debug(str_debug)
            if len(list_results) != 0:
                list_results, list_resultnumbers = renumber_testcase_action(list_results, "RESULTS")
            else:
                # print(char_debug_print * outline_left, "NO RESULTS------> len(list_results)", len(list_results))
                str_debug = f"{char_debug_print * outline_left} NO RESULTS------> len(list_results) {len(list_results)}"
                logger.debug(str_debug)
                list_results = []
                list_resultnumbers = []
#            list_resultnumbers = list_results[:]
            # print(char_debug_print * outline_left, "R E N U M B E R  +++++++++>", row, len(list_steps), len(list_results))
            str_debug = f"{char_debug_print * outline_left} R E N U M B E R  +++++++++> {row} \
                                {len(list_steps)} {len(list_results)}"
            logger.debug(str_debug)
            # print(char_debug_print * outline_left, "------ R E S U L T S ---------- list_results", list_results)
            str_debug = f"{char_debug_print * outline_left} ------ R E S U L T S ---------- list_results {list_results}"
            logger.debug(str_debug)
            # print(char_debug_print * outline_left, "------ R E S U L T S ---------- list_resultnumbers", list_resultnumbers)
            str_debug = f"{char_debug_print * outline_left} ------ R E S U L T S ---------- list_resultnumbers {list_resultnumbers}"
            logger.debug(str_debug)
            cs = 0
            without_number = 0  # steps and numbers without numbers
            for x in list_steps:
                cs += 1
                # print(char_debug_print * outline_left, cs, x)
                str_debug = f"{char_debug_print * outline_left} {cs} {x}"
                logger.debug(str_debug)
                if "Precondition" in x:
                    precondition = x
                    if "." in x:
                        split_precondition_steps = x.split('.')  # Check first point
                        # print(char_debug_print * outline_left, "YYYYYYYY split_precondition_steps", split_precondition_steps)
                        str_debug = f"{char_debug_print * outline_left} YYYYYYYY split_precondition_steps {split_precondition_steps}"
                        logger.debug(str_debug)
                        # print(char_debug_print * outline_left,
                        #      "YYYYYYYY len(split_precondition_steps)", len(split_precondition_steps))
                        str_debug = f"{char_debug_print * outline_left} YYYYYYYY len(split_precondition_steps) {len(split_precondition_steps)}"
                        logger.debug(str_debug)
                        if (len(split_precondition_steps) > 0):
                            for i in range(0, len(split_precondition_steps)):
                                # print(char_debug_print * outline_left, "i", i)
                                str_debug = f"{char_debug_print * outline_left} i {i}"
                                logger.debug(str_debug)
                                # print(char_debug_print * outline_left,
                                #      "split_precondition_steps[i] :", split_precondition_steps[i], "YYYYYYYYYYYYY")
                                str_debug = f"{char_debug_print * outline_left} split_precondition_steps[i]  \
                                                {split_precondition_steps[i]} YYYYYYYYYYYYY"
                                logger.debug(str_debug)
                                if "Precondition" in split_precondition_steps[i]:
                                    # print(char_debug_print * outline_left,
                                    #      "Precondition Catch", split_precondition_steps[i])
                                    str_debug = f"{char_debug_print * outline_left} Precondition Catch  \
                                                {split_precondition_steps[i]}"
                                    logger.debug(str_debug)
                                    x = split_precondition_steps[i].strip()
                                    break
                    write_col = 4  # Steps : read column F(6); Preconditions : write column E(number_of_spaces)
                    # print(char_debug_print * outline_left, "Precondition ", cs, x.strip())
                    str_debug = f"{char_debug_print * outline_left} Precondition {cs} {x.strip()}"
                    logger.debug(str_debug)
                    # precondition = tab.cell(row=row_start, column=write_col)
                    # precondition.value = x.strip()
                    worksheet.write(row_start, write_col, x.strip())
                    continue
                if len(x.strip()) == 0:
                    # print(char_debug_print * outline_left, "Steps Blank", cs, x)
                    str_debug = f"{char_debug_print * outline_left} Steps Blank {cs} {x}"
                    logger.debug(str_debug)
                    continue
#                    for i in range(9):
#                        stepnr = i + 1
                regex_stepnr = re.compile(r"^[0-9][0-9]?")
                list_stepnr = regex_stepnr.findall(x)
                if len(list_stepnr) == 1:
                    int_stepnr = int(list_stepnr[0])
                    # print(char_debug_print * outline_left, "int_stepnr", type(int_stepnr), int_stepnr)
                    str_debug = f"{char_debug_print * outline_left} int_stepnr {type(int_stepnr)} {int_stepnr}"
                    logger.debug(str_debug)
                    # print(char_debug_print * outline_left, "split before", row_start)
                    str_debug = f"{char_debug_print * outline_left} split before {row_start}"
                    logger.debug(str_debug)
                    row_start = split_steps(int_stepnr, row_start, x, list_results,
                                            list_steps, list_stepnumbers, list_resultnumbers)
                    # print(char_debug_print * outline_left, "split after", row_start)
                    str_debug = f"{char_debug_print * outline_left} split after {row_start}"
                    logger.debug(str_debug)
                    row_start += 1
                else:  # Steps and Results without numbers
                    if without_number == 0:
                        steps_start = row_start
                        without_number = 1
                        # print(char_debug_print * outline_left, "step_start", steps_start)
                        str_debug = f"{char_debug_print * outline_left} step_start {steps_start}"
                        logger.debug(str_debug)
                    regex_steps = re.compile(r"^[a-zA-Z][a-zA-Z]", re.MULTILINE)
                    list_stepxx = regex_steps.findall(test_steps)
                    # print(char_debug_print * outline_left, "len(list_stepxx)", len(list_stepxx), list_stepxx)
                    str_debug = f"{char_debug_print * outline_left} len(list_stepxx) {len(list_stepxx)} {list_stepxx}"
                    logger.debug(str_debug)
                    # print(char_debug_print * outline_left, "+" * 40)
                    str_debug = f"{char_debug_print * outline_left} {'+' * 40}"
                    logger.debug(str_debug)
                    # print(char_debug_print * outline_left, "test_steps", len(test_steps), test_steps)
                    str_debug = f"{char_debug_print * outline_left} test_steps {len(test_steps)} {test_steps}"
                    logger.debug(str_debug)
                    # print(char_debug_print * outline_left, "+" * 40)
                    str_debug = f"{char_debug_print * outline_left} {'+' * 40}"
                    logger.debug(str_debug)
                    # print(char_debug_print * outline_left, "test_results", test_results)
                    str_debug = f"{char_debug_print * outline_left} test_results {test_results}"
                    logger.debug(str_debug)
                    if (len(list_stepxx) > 0):
                        split_test_steps = test_steps.split('\n')
                        # print(char_debug_print * outline_left, "XXXXXXXXX split_test_steps", split_test_steps)
                        str_debug = f"{char_debug_print * outline_left} XXXXXXXXX split_test_steps {split_test_steps}"
                        logger.debug(str_debug)
                        # print(char_debug_print * outline_left, "XXXXXXXXX len(split_test_steps)", len(split_test_steps))
                        str_debug = f"{char_debug_print * outline_left} XXXXXXXXX len(split_test_steps) {len(split_test_steps)}"
                        logger.debug(str_debug)
                        c = len(list_stepxx) - 1
                        # print(char_debug_print * outline_left, "c: ", c)
                        str_debug = f"{char_debug_print * outline_left} c {c}"
                        logger.debug(str_debug)
                        for i in range(1, len(split_test_steps)):
                            # print(char_debug_print * outline_left, "i", i)
                            str_debug = f"{char_debug_print * outline_left} i {i}"
                            logger.debug(str_debug)
                            # print(char_debug_print * outline_left,
                            #      "split_test_steps[1] :", split_test_steps[1], "XXXXXXXXXXXX")
                            str_debug = f"{char_debug_print * outline_left} split_test_steps[1] {split_test_steps[1]} XXXXXXXXXXXX"
                            logger.debug(str_debug)
                            # print("split", split_test_steps[i])
                            # print(char_debug_print * outline_left, "list_stepxx[c] :", list_stepxx[c], "XXXXXXXXXXXX")
                            str_debug = f"{char_debug_print * outline_left} list_stepxx[c] {list_stepxx[c]} XXXXXXXXXXXX"
                            logger.debug(str_debug)
                            if split_test_steps[i].startswith(list_stepxx[c]):
                                # print(char_debug_print * outline_left, "Catch", split_test_steps[i])
                                str_debug = f"{char_debug_print * outline_left} Catch {split_test_steps[i]}"
                                logger.debug(str_debug)
                                test_steps = split_test_steps[i]
                                test_results = ""
                                break
                        stepxx = test_steps
                        resultxx = test_results
                        # print(char_debug_print * outline_left, "stepxx", stepxx)
                        str_debug = f"{char_debug_print * outline_left} stepxx {stepxx}"
                        logger.debug(str_debug)
                        # if ("Precondition" not in stepxx):
                        flag_debug_matches = 7
                        write_to_excel_template(stepxx, resultxx, write_row_start,
                                                flag_remove_number, flag_debug_matches)
                        row_start += 1
        else:  # Steps are blank and results are present
            # print(char_debug_print * outline_left, "STEPS ARE BLANK AND RESULTS ARE PRESENT")
            str_debug = f"{char_debug_print * outline_left} STEPS ARE BLANK AND RESULTS ARE PRESENT"
            logger.debug(str_debug)

    # print(char_debug_print * outline_left, "check_steps_and_results OUT parameter row_star", row_start)
    str_debug = f"{char_debug_print * outline_left} check_steps_and_results OUT parameter row_start {row_start}"
    logger.debug(str_debug)
    # print(char_routine_stop * outline_left,
    #      dict_routines["check_steps_and_results"][0], char_routine_stop * outline_right)
    str_debug = f"{char_routine_stop * outline_left} {dict_routines['check_steps_and_results'][0]} {char_routine_stop * outline_right} "
    logger.debug(str_debug)

    return row_start


# ------------------MAIN -------------------------


# add filemode="w" to overwrite
# logging.basicConfig(filename="import_testcases.log",
#                    level=logging.DEBUG,
#                    format='%(levelname)-8s %(message)s',
#                    filemode='w')

logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s %(levelname)-8s [%(filename)s:%(lineno)d]: %(message)s',
                    datefmt='%d-%m-%Y %H:%M:%S',
                    filename='import_testcases.log',
                    filemode='w')
logger = logging.getLogger('import_testcases_logger')

# logging.debug("This is a debug message")
# logging.info("Informational message")
# logging.error("An error has happened!")


char_routine_start = ">"
char_routine_stop = "<"
char_debug_print = "."
number_of_spaces = 2
number_char_divider = 110
flag_remove_number = 0  # Remove step or result number
flag_debug_matches = 1  # Print debug info function write_step_matches_result
dict_routines = {
    "main": ["main", 0, 4],
    "select_excel_tab": ["select_excel_tab", 1, 16],
    "write_header": ["write_header", 1, 12],
    "check_steps_and_results": ["check_steps_and_results", 1, 22],
    "renumber_testcase_action": ["renumber_testcase_action", 2, 24],
    "split_steps": ["split_steps", 2, 11],
    "write_to_excel_template": ["write_to_excel_template", 2, 23],
}
# Build regex dictionary
dict_regex = {}
keys = range(100)
regex_string = 're.compile(r"^'
for i in keys:
    dict_regex[i] = regex_string + str(i) + '")'

outline_left, outline_right = calculate_outline(dict_routines["main"][0], dict_routines["main"][1])
str_debug = f"{char_routine_start * outline_left} {dict_routines['main'][0]} {char_routine_start * outline_right}"
logger.debug(str_debug)
# print(char_routine_start * outline_left, dict_routines["main"][0], char_routine_start * outline_right)

list_col_access = [2, 3, 4, 5, 6]
list_col_copy = [2, 3, 4, 5]
list_col_split = [6, 7]
list_steps = []
list_header1 = ["Test Suite", "Test Cases"]
list_header2 = ["Name", "Details", "Name", "Summary", "Preconditions", "Test Execution Type",
                "Importance", "Steps", "Expected Results", "Step Execution Type", "Requirements"]
write_row_start = 2
write_row_stop = 2
list_TC_topics = ["Install & Uninstall", "First Run & Registration", "Settings",
                  "New Label and Open Label", "Printers & Consumables", "Label Editing",
                  "Import Data", "Text Based Objects", "Graphic Based Objects"]
tabnr = select_excel_tab()
inputfile = "Reviewed Test cases_DCD_Updated_19062018.xlsx"
wb = openpyxl.load_workbook(inputfile)
wb_template = xlsxwriter.Workbook('TC_template.xlsx')
worksheet = wb_template.add_worksheet("TestCases")
# wb = openpyxl.load_workbook('Test cases_DCD_FN.xlsx')
for sheet in wb.sheetnames:
    if sheet == "Totaal reviewed cases":
        continue
    if sheet == "Import Data":
        # print(char_debug_print * outline_left, "SKIP TAB NO DATA----------------->", sheet, list_TC_topics[tabnr])
        str_debug = f"{char_debug_print * outline_left} SKIP TAB NO DATA-----------------> {sheet} {list_TC_topics[tabnr]}"
        logger.debug(str_debug)
        continue
    if sheet == "Install & Uninstall":
        write_header()
#        continue  # discard TOC
    if sheet != list_TC_topics[tabnr]:
        # print(char_debug_print * outline_left, "SKIP TAB", sheet, list_TC_topics[tabnr])
        str_debug = f"{char_debug_print * outline_left} SKIP TAB {sheet} {list_TC_topics[tabnr]}"
        logger.debug(str_debug)
        continue
    if sheet in list_TC_topics:
        # print(char_debug_print * outline_left, "D008", sheet, list_TC_topics[tabnr])
        str_debug = f"{char_debug_print * outline_left} {sheet} {list_TC_topics[tabnr]}"
        logger.debug(str_debug)
        tab = wb[sheet]
        for row in range(2, tab.max_row + 1):
            #  write_row_start += 1
            for col in range(1, tab.max_column + 1):
                if col not in list_col_access:
                    e = ""
                    continue
                e = tab.cell(row=row, column=col)
                if col in list_col_copy:
                    if e.value is not None:
                        if col == 2:  # Traceablity ID : read column B (02); Requirements : write column K(11)
                            write_col = 10
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 3:  # Functionality : read column C (03); Name : write column A(01)
                            write_col = 0
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 4:  # Task : read column D (04); Name : write column C(03)
                            write_col = 2
                            e = tab.cell(row=row, column=col)
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 5:  # Description : read column E (05); Summary : write column D(04)
                            write_col = 3
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 8:  # Priority : read column H (8); Importance : write column G(7)
                            write_col = 6
                            worksheet.write(write_row_start, write_col, e.value)
                if col in list_col_split:
                    # print(char_debug_print * outline_left, "D009", "row_start", write_row_start)
                    str_debug = f"{char_debug_print * outline_left} row_start {write_row_start}"
                    logger.debug(str_debug)
                    write_row_start = check_steps_and_results(row, col, write_row_start)
wb_template.close()  # save the workbook
wb.close()
# print(char_routine_stop * outline_left, dict_routines["main"][0], char_routine_stop * outline_right)
str_debug = f"{char_routine_start * outline_left} {dict_routines['main'][0]} {char_routine_stop * outline_right}"
logger.debug(str_debug)
