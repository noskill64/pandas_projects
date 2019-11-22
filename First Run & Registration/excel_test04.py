import openpyxl
import re
import sys
import xlsxwriter
from openpyxl.styles import PatternFill


def split_steps(nr, rijnr, step, list_r):
    print("split_steps IN parameter nr ", nr)
    print("split_steps IN parameter rijnr ", rijnr)
    print("split_steps IN parameter step ", step)
    print("split_steps IN parameter list_r", list_r)
    write_row_start = rijnr
    write_row_stop = rijnr
    if nr == 1:
        print("steps_regex1")
        regex_sxx = re.compile(r"^1")
    if nr == 2:
        print("steps_regex2")
        regex_sxx = re.compile(r"^2")
    if nr == 3:
        print("steps_regex3")
        regex_sxx = re.compile(r"^3")
    if nr == 4:
        print("steps_regex4")
        regex_sxx = re.compile(r"^4")
    if nr == 5:
        print("steps_regex5")
        regex_sxx = re.compile(r"^5")
    if nr == 6:
        print("steps_regex6")
        regex_sxx = re.compile(r"^6")
    if nr == 7:
        print("steps_regex7")
        regex_sxx = re.compile(r"^7")
    if nr == 8:
        print("steps_regex8")
        regex_sxx = re.compile(r"^8")
    if nr == 9:
        print("steps_regex9")
        regex_sxx = re.compile(r"^9")
    if nr == 10:
        print("steps_regex10")
        regex_sxx = re.compile(r"^10")
    if nr == 11:
        print("steps_regex11")
        regex_sxx = re.compile(r"^11")
    if nr == 12:
        print("steps_regex12")
        regex_sxx = re.compile(r"^12")
    if nr == 13:
        print("steps_regex13")
        regex_sxx = re.compile(r"^13")
    if nr == 14:
        print("steps_regex14")
        regex_sxx = re.compile(r"^14")
    if nr == 15:
        print("steps_regex15")
        regex_sxx = re.compile(r"^15")
    if nr == 16:
        print("steps_regex16")
        regex_sxx = re.compile(r"^16")
    if nr == 17:
        print("steps_regex17")
        regex_sxx = re.compile(r"^17")
    if nr == 18:
        print("steps_regex18")
        regex_sxx = re.compile(r"^18")
    if nr == 19:
        print("steps_regex19")
        regex_sxx = re.compile(r"^19")
    if nr == 20:
        print("steps_regex20")
        regex_sxx = re.compile(r"^20")

    list_sxx = regex_sxx.findall(step)
    if len(list_sxx) > 0:  # Step01
        stepxx = step
        resultxx = ""
        if nr == 1:
            print("results_regex1")
            regex_rxx = re.compile(r"^1")
        if nr == 2:
            print("results_regex2")
            regex_rxx = re.compile(r"^2")
        if nr == 3:
            print("results_regex3")
            regex_rxx = re.compile(r"^3")
        if nr == 4:
            print("results_regex4")
            regex_rxx = re.compile(r"^4")
        if nr == 5:
            print("results_regex5")
            regex_rxx = re.compile(r"^5")
        if nr == 6:
            print("results_regex6")
            regex_rxx = re.compile(r"^6")
        if nr == 7:
            print("results_regex7")
            regex_rxx = re.compile(r"^7")
        if nr == 8:
            print("results_regex8")
            regex_rxx = re.compile(r"^8")
        if nr == 9:
            print("results_regex9")
            regex_rxx = re.compile(r"^9")
        if nr == 10:
            print("results_regex10")
            regex_rxx = re.compile(r"^10")
        if nr == 11:
            print("results_regex11")
            regex_rxx = re.compile(r"^11")
        if nr == 12:
            print("results_regex12")
            regex_rxx = re.compile(r"^12")
        if nr == 13:
            print("results_regex13")
            regex_rxx = re.compile(r"^13")
        if nr == 14:
            print("results_regex14")
            regex_rxx = re.compile(r"^14")
        if nr == 15:
            print("results_regex15")
            regex_rxx = re.compile(r"^15")
        if nr == 16:
            print("results_regex16")
            regex_rxx = re.compile(r"^16")
        if nr == 17:
            print("results_regex17")
            regex_rxx = re.compile(r"^17")
        if nr == 18:
            print("results_regex18")
            regex_rxx = re.compile(r"^18")
        if nr == 19:
            print("results_regex19")
            regex_rxx = re.compile(r"^19")
        if nr == 20:
            print("results_regex20")
            regex_rxx = re.compile(r"^20")

        match_rxx = 0  # matching pair step - result
        match_rxx_char = 0  # step is present - result has second line
        match_rxx_blank = 0  # steps is present - result has blank line
        match_position = 0
        match_row = 0
        count = 0
        for y in list_r:
            if (match_rxx == 1) and (len(y.strip()) == 0):
                print("<<<skip blank y")
                break
            count += 1
            print("===", count, y)
            list_rxx = regex_rxx.findall(y)
            if len(list_rxx) > 0:  # Matching steps 1
                match_position = count
                match_rxx = 1
                resultxx = y
                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                if "." in stepxx:
                    point_pos = stepxx.find(".")
                    if (point_pos > 0) and (point_pos < 4):  # Take first point after number, not ending point
                        stepxx = stepxx[point_pos + 1:]
                worksheet.write(write_row_start, write_col, stepxx)
                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                if "." in resultxx:
                    point_pos = resultxx.find(".")
                    if (point_pos > 0) and (point_pos < 4):  # Take first point after number, not ending point
                        resultxx = resultxx[point_pos + 1:]
                match_row = write_row_start
                worksheet.write(write_row_start, write_col, resultxx)
                print("result ", nr, stepxx, resultxx, write_row_start)
            else:
                regex_rxx_char = re.compile(r"^[a-zA-Z]")
                list_rxx_char = regex_rxx_char.findall(y)
                if (count == match_position + 1) and (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    step_blank = ""
                    resultxx = y
                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                    worksheet.write(write_row_stop, write_col, step_blank)
                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                    if "." in resultxx:
                        point_pos = resultxx.find(".")
                        if (point_pos > 0) and (point_pos < 4):  # Take first point after number, not ending point
                            resultxx = resultxx[point_pos + 1:]
                    worksheet.write(write_row_stop, write_col, resultxx)  # schrijf lijn + 1
                    print("result", nr, "NOT NULL", step_blank, resultxx, write_row_stop)
                    continue  # stop zoeken naar results
                if (match_rxx_char == 1) and (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    step_blank = ""
                    resultxx = y
                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                    worksheet.write(write_row_stop, write_col, step_blank)
                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                    if "." in resultxx:
                        point_pos = resultxx.find(".")
                        if (point_pos > 0) and (point_pos < 4):  # Take first point after number, not ending point
                            resultxx = resultxx[point_pos + 1:]
                    worksheet.write(write_row_stop, write_col, resultxx)  # schrijf lijn + 1
                    print("NEXT ABC RESULTS", nr, step_blank, resultxx, write_row_stop)
                    continue
                if (match_rxx_blank == 0) and (match_rxx == 0) and (len(y.strip()) == 0):  # result01 = blank
                    match_rxx_blank = 1
                    resultxx = ""
                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                    if "." in stepxx:
                        point_pos = stepxx.find(".")
                        if (point_pos > 0) and (point_pos < 4):  # Take first point after number, not ending point
                            stepxx = stepxx[point_pos + 1:]
                    worksheet.write(write_row_start, write_col, stepxx)
                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                    worksheet.write(write_row_start, write_col, resultxx)
                    print("result", nr, " BLANK", stepxx, resultxx, write_row_start)
                    continue
        if list_r == "":
            resultxx = ""
            write_col = 7  # Steps : read column F(6); Steps : write column H(8)
            if "." in stepxx:
                point_pos = stepxx.find(".")
                if (point_pos > 0) and (point_pos < 4):  # Take first point after number, not ending point
                    stepxx = stepxx[point_pos + 1:]
            worksheet.write(write_row_start, write_col, stepxx)
            write_col = 8  # Steps : read column F(6); Steps : write column I(9)
            worksheet.write(write_row_start, write_col, resultxx)
            print("MISSING RESULTS", nr, " BLANK", stepxx, resultxx, write_row_start)
        elif (match_rxx == 0):
            resultxx = ""
            write_col = 7  # Steps : read column F(6); Steps : write column H(8)
            if "." in stepxx:
                point_pos = stepxx.find(".")
                if (point_pos > 0) and (point_pos < 4):  # Take first point after number, not ending point
                    stepxx = stepxx[point_pos + 1:]
            worksheet.write(write_row_start, write_col, stepxx)
            write_col = 8  # Steps : read column F(6); Steps : write column I(9)
            worksheet.write(write_row_start, write_col, resultxx)
            print("WHAT THE FUCK", list_r, "$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("OUT parameter", write_row_stop)
    return write_row_stop


def check_steps_and_results(read_row, read_col, row_start):
    print("check_steps_and_results IN parameter read_row ", read_row)
    print("check_steps_and_results IN parameter read_col ", read_col)
    print("check_steps_and_results IN parameter row_start ", row_start)

    row_stop = row_start
    e = tab.cell(row=read_row, column=read_col)
    if e.value is not None:

        s = tab.cell(row=read_row, column=read_col)
        r = tab.cell(row=read_row, column=read_col + 1)
        if (s.value is not None):  # Steps are present
            test_steps = s.value
            list_steps = test_steps.splitlines()
            if (r.value is not None):  # Results are present
                test_results = r.value
                list_results = test_results.splitlines()
            else:
                test_results = ""
                list_results = ""
#            if len(list_steps) != len(list_results):
            print("++++++++++++++++++++++>", row, len(list_steps), len(list_results))
            cs = 0
            step01 = step02 = step03 = step04 = step05 = ""
            step06 = step07 = step08 = step09 = ""
            without_number = 0  # steps and numbers without numbers
            for x in list_steps:
                cs += 1
                print(cs, x)
                if "Precondition" in x:
                    precondition = x
                    if "." in x:
                        split_precondition_steps = x.split('.')  # Check first point
                        print("YYYYYYYY split_precondition_steps", split_precondition_steps)
                        print("YYYYYYYY len(split_precondition_steps)", len(split_precondition_steps))
                        if (len(split_precondition_steps) > 0):
                            for i in range(0, len(split_precondition_steps)):
                                print("i", i)
                                print("split_precondition_steps[i] :", split_precondition_steps[i], "YYYYYYYYYYYYY")
                                if "Precondition" in split_precondition_steps[i]:
                                    print("Precondition Catch", split_precondition_steps[i])
                                    x = split_precondition_steps[i].strip()
                                    break
                    write_col = 4  # Steps : read column F(6); Preconditions : write column E(5)
                    print("Precondition ", cs, x.strip())
                    # precondition = tab.cell(row=row_start, column=write_col)
                    # precondition.value = x.strip()
                    worksheet.write(row_start, write_col, x.strip())
                    continue
                if len(x.strip()) == 0:
                    print("Steps Blank", cs, x)
                    continue
#                    for i in range(9):
#                        stepnr = i + 1
                regex_stepnr = re.compile(r"^[0-9][0-9]?")
                list_stepnr = regex_stepnr.findall(x)
                if len(list_stepnr) == 1:
                    int_stepnr = int(list_stepnr[0])
                    print("int_stepnr", type(int_stepnr), int_stepnr)
                    print("split before", row_start)
                    row_start = split_steps(int_stepnr, row_start, x, list_results)
                    print("split after", row_start)
                    row_start += 1
                else:  # Steps and Results without numbers
                    if without_number == 0:
                        steps_start = row_start
                        without_number = 1
                        print("step_start", steps_start)
                    regex_steps = re.compile(r"^[a-zA-Z][a-zA-Z]", re.MULTILINE)
                    list_stepxx = regex_steps.findall(test_steps)
                    print("len(list_stepxx)", len(list_stepxx), list_stepxx)
                    print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                    print("test_steps", len(test_steps), test_steps)
                    print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                    print("test_results", test_results)
                    if (len(list_stepxx) > 0):
                        split_test_steps = test_steps.split('\n')
                        print("XXXXXXXXX split_test_steps", split_test_steps)
                        print("XXXXXXXXX len(split_test_steps)", len(split_test_steps))
                        c = len(list_stepxx) - 1
                        print("c: ", c)
                        for i in range(1, len(split_test_steps)):
                            print("i", i)
                            print("split_test_steps[1] :", split_test_steps[1], "XXXXXXXXXXXX")
                            # print("split", split_test_steps[i])
                            print("list_stepxx[c] :", list_stepxx[c], "XXXXXXXXXXXX")
                            if split_test_steps[i].startswith(list_stepxx[c]):
                                print("Catch", split_test_steps[i])
                                test_steps = split_test_steps[i]
                                test_results = ""
                                break
                        stepxx = test_steps
                        resultxx = test_results
                        print("stepxx", stepxx)
                        # if ("Precondition" not in stepxx):
                        write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                        worksheet.write(row_start, write_col, stepxx)
                        write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                        worksheet.write(row_start, write_col, resultxx)
                        print("STEPS WITHOUT NUMBERS ???? ", "1", stepxx, resultxx, row_start)
                        row_start += 1
        else:  # Steps are blank and results are present
            print("STEPS ARE BLANK AND RESULTS ARE PRESENT")
    return row_start


# ------------------MAIN -------------------------


list_col_access = [2, 3, 4, 5, 6]
list_col_copy = [2, 3, 4, 5]
list_col_split = [6, 7]
list_steps = []
list_header1 = ["Test Suite", "Test Cases"]
list_header2 = ["Name", "Details", "Name", "Summary", "Preconditions", "Test Execution Type",
                "Importance", "Steps", "Expected Results", "Step Execution Type", "Requirements"]
write_row_start = 2
write_row_stop = 2
list_TC_topics = ["Install & Uninstall", "First Run & Registration", "Settings",
                  "New Label and Open Label", "Connect & Recognize Printers",
                  "Printers & Consumables", "Text Based Objects", "Graphic Based Objects"]
print("Number of arguments: ", len(sys.argv), "arguments")
if len(sys.argv) > 1:
    print("Argument list: ", str(sys.argv), type(sys.argv))
    if len(sys.argv) == 2:
        if sys.argv[1].isnumeric():
            print("isnumeric", sys.argv[1])
            number_arg = int(sys.argv[1])
            if (number_arg > 0) and (number_arg < 9):
                number_arg -= 1
                print(number_arg, list_TC_topics[number_arg])
    else:
        for number_arg in range(1, len(sys.argv)):
            print(number_arg, list_TC_topics[number_arg])

wb = openpyxl.load_workbook('Reviewed Test cases_DCD_Updated_09042018.xlsx')
wb_template = xlsxwriter.Workbook('TC_template.xlsx')
worksheet = wb_template.add_worksheet("TestCases")
# wb = openpyxl.load_workbook('Test cases_DCD_FN.xlsx')
for sheet in wb.sheetnames:
    if sheet == "Totaal reviewed cases":
        continue
    if sheet == "Table of Contents":
        #  Print header 1
        header1_row = 0
        header1_col1 = 0
        header1_col2 = 2
        # Create a format to use in the merged range.
        merge_format = wb_template.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '808080'})
        # Merge 2 cells.
        print("Header1", header1_row, header1_col1, list_header1[0])  # Header 1 Column Test Suite
        worksheet.write(header1_row, header1_col1, list_header1[0])
        worksheet.merge_range('A1:B1', list_header1[0], merge_format)
        # Merge 9 cells.
        print("Header1", header1_row, header1_col2, list_header1[1])  # Header 1 Column Test Cases
        worksheet.write(header1_row, header1_col2, list_header1[1])
        worksheet.merge_range('C1:K1', list_header1[1], merge_format)  # format gray text
        #  print header 2
        header2_row = 1
        header2_max_col = 11
        for header2_col in range(header2_max_col):
            print("Header2", header2_row, header2_col, list_header2[header2_col])
        #    worksheet.cell(header2_row, header2_col).fill = PatternFill(fgColor='808080', fill_type='solid')
            worksheet.write(header2_row, header2_col, list_header2[header2_col], merge_format)
        #    my_fill = PatternFill(fgColor='FFFFFF', bgColor='FFFFFF', fill_type='solid')
        #    worksheet['A2'].fill = my_fill
        continue  # discard TOC
    if sheet != list_TC_topics[number_arg]:
        continue
    if sheet in list_TC_topics:
        print("+++++++++++++++++++++++++++++++++++++", "START OF ", sheet, "+++++++++++++++++++++++++++++++++++++")
        list_steps = []
        tab = wb[sheet]
        for row in range(2, tab.max_row + 1):
            #  write_row_start += 1
            for col in range(1, tab.max_column + 1):
                if col not in list_col_access:
                    e = ""
                    continue
                e = tab.cell(row=row, column=col)
                if col in list_col_copy:
                    if e.value is not None:
                        if col == 2:  # Traceablity ID : read column B (02); Requirements : write column K(11)
                            write_col = 10
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 3:  # Functionality : read column C (03); Name : write column A(01)
                            write_col = 0
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 4:  # Task : read column D (04); Name : write column C(03)
                            write_col = 2
                            e = tab.cell(row=row, column=col)
                            print("+++++Name++++++", e.value, "row:", row, "col:", col)
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 5:  # Description : read column E (05); Summary : write column D(04)
                            write_col = 3
                            print("Summary", e.value)
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 8:  # Priority : read column H (8); Importance : write column G(7)
                            write_col = 6
                            worksheet.write(write_row_start, write_col, e.value)
                if col in list_col_split:
                    print("main_write_row_start", write_row_start)
                    write_row_start = check_steps_and_results(row, col, write_row_start)
wb_template.close()  # save the workbook
wb.close()
