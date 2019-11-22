import openpyxl
import re
import sys
import xlsxwriter
from openpyxl.styles import PatternFill


def renumber_testcase_action(list_action, results):
    j = 0
    for i in range(0, len(list_action)):
        if "Precondition" in list_action[i]:
            continue
        print("renum_step", j, list_action[j])
        j += 1
    for k in range(0, len(results)):
        print("renum_result", k, results[k])
    letter_counter = 0
    flag_action_number = ""
    no_action = 0
    list_alphabet = ["a.", "b.", "c.", "d.", "f.", "g.", "h.", "i.", "j.", "k.", "l.", "m.", "n.", "o.",
                     "p.", "q.", "r.", "s.", "t.", "u.", "v.", "w.", "x.", "y.", "z."]
    for i in range(0, len(list_action)):
        if "Precondition" in list_action[i]:
            continue
        print("Action", list_action[i])
        if "." in list_action[i]:  # point present
            print("punt aanwezig", list_action[i])
            regex_actionnr = re.compile(r"^[0-9][0-9]?")
            list_actionnr = regex_actionnr.findall(list_action[i])
            if len(list_actionnr) == 1:
                split_number_action = list_action[i].split('.')  # Check first point
                action_number = split_number_action[0]
                letter_counter = 0
                no_action = 1
                flag_action_number = action_number
                print("Action number", action_number, split_number_action)
            else:
                no_action = 1
                list_action[i] = "1." + list_action[i]
                print("point without number", list_action[i])
        else:  # point not present
            print("punt NIET aanwezig", list_action[i])
            if len(list_action[i].strip()) == 0:  # Blanks
                if no_action == 0:  # skip first blank step line
                    no_action = 1
                    action_number = "1"
                list_action[i] = "BLANK"
                if flag_action_number != action_number:
                    flag_action_number = action_number
                    letter_counter = 0
                    action_number_next = action_number + list_alphabet[letter_counter]
                else:
                    action_number_next = action_number + list_alphabet[letter_counter]
                    letter_counter += 1
                list_action[i] = action_number_next + list_action[i]
                print("Steps Blank", list_action[i])
            else:  # Starts with letter, no step number
                print("Starts with letter, no step number")
                regex_action = re.compile(r"^[a-zA-Z][a-zA-Z]", re.MULTILINE)
                list_char_detect = regex_action.findall(list_action[i])
                if (len(list_char_detect) > 0):
                    if flag_action_number != action_number:
                        flag_action_number = action_number
                        letter_counter = 0
                        action_number_next = action_number + list_alphabet[letter_counter]
                    else:
                        action_number_next = action_number + list_alphabet[letter_counter]
                        letter_counter += 1
                    list_action[i] = action_number_next + list_action[i]
                    print("Action char", list_action[i])
    return list_action


def select_excel_tab():
    print("Number of arguments: ", len(sys.argv), "arguments")
    if len(sys.argv) > 1:
        print("Argument list: ", str(sys.argv), type(sys.argv))
        if len(sys.argv) == 2:
            if sys.argv[1].isnumeric():
                print("isnumeric", sys.argv[1])
                number_arg = int(sys.argv[1])
                if (number_arg > 0) and (number_arg < 9):
                    number_arg -= 1
                    print(number_arg, list_TC_topics[number_arg])
        else:
            for number_arg in range(1, len(sys.argv)):
                print(number_arg, list_TC_topics[number_arg])
    return number_arg


def write_header():
    #  Print header 1
    header1_row = 0
    header1_col1 = 0
    header1_col2 = 2
    # Create a format to use in the merged range.
    merge_format = wb_template.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '808080'})
    # Merge 2 cells.
    print("Header1", header1_row, header1_col1, list_header1[0])  # Header 1 Column Test Suite
    worksheet.write(header1_row, header1_col1, list_header1[0])
    worksheet.merge_range('A1:B1', list_header1[0], merge_format)
    # Merge 9 cells.
    print("Header1", header1_row, header1_col2, list_header1[1])  # Header 1 Column Test Cases
    worksheet.write(header1_row, header1_col2, list_header1[1])
    worksheet.merge_range('C1:K1', list_header1[1], merge_format)  # format gray text
    #  print header 2
    header2_row = 1
    header2_max_col = 11
    for header2_col in range(header2_max_col):
        print("Header2", header2_row, header2_col, list_header2[header2_col])
        #    worksheet.cell(header2_row, header2_col).fill = PatternFill(fgColor='808080', fill_type='solid')
        worksheet.write(header2_row, header2_col, list_header2[header2_col], merge_format)
        #    my_fill = PatternFill(fgColor='FFFFFF', bgColor='FFFFFF', fill_type='solid')
        #    worksheet['A2'].fill = my_fill


def write_to_excel_template(step, result, row, flag_rn, flag_dm):

    col = 7  # Steps : read column F(6); Steps : write column H(8)
    if flag_remove_number == 1:
        step = remove_numbers(step)
    worksheet.write(row, col, step)
    col = 8  # Steps : read column F(6); Steps : write column I(9)
    if flag_remove_number == 1:
        result = remove_numbers(result)
    worksheet.write(row, col, result)
    if flag_dm == 1:
        print("MATCH", step, result, row)
    if flag_dm == 2:
        print("STEP=BLANK, RESULT  WITHOUT NUMBER", step, result, row)
    if flag_dm == 3:
        print("NEXT ABC RESULTS", step, result, row)
    if flag_dm == 4:
        print("RESULT = BLANK", step, result, row)
    if flag_dm == 5:
        print("MISSING RESULTS", step, result, row)
    if flag_dm == 6:
        print("WHAT THE FUCK", step, result, row, "$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    if flag_dm == 7:
        print("STEPS WITHOUT NUMBERS ???? ", step, result, row)


def select_regex_string(number):

    regex_string = ""
    if number == 1:
        print("regex1")
        regex_string = re.compile(r"^1")
    if number == 2:
        print("regex2")
        regex_string = re.compile(r"^2")
    if number == 3:
        print("regex3")
        regex_string = re.compile(r"^3")
    if number == 4:
        print("regex4")
        regex_string = re.compile(r"^4")
    if number == 5:
        print("regex5")
        regex_string = re.compile(r"^5")
    if number == 6:
        print("regex6")
        regex_string = re.compile(r"^6")
    if number == 7:
        print("regex7")
        regex_string = re.compile(r"^7")
    if number == 8:
        print("regex8")
        regex_string = re.compile(r"^8")
    if number == 9:
        print("regex9")
        regex_string = re.compile(r"^9")
    if number == 10:
        print("regex10")
        regex_string = re.compile(r"^10")
    if number == 11:
        print("regex11")
        regex_string = re.compile(r"^11")
    if number == 12:
        print("regex12")
        regex_string = re.compile(r"^12")
    if number == 13:
        print("regex13")
        regex_string = re.compile(r"^13")
    if number == 14:
        print("regex14")
        regex_string = re.compile(r"^14")
    if number == 15:
        print("regex15")
        regex_string = re.compile(r"^15")
    if number == 16:
        print("regex16")
        regex_string = re.compile(r"^16")
    if number == 17:
        print("regex17")
        regex_string = re.compile(r"^17")
    if number == 18:
        print("regex18")
        regex_string = re.compile(r"^18")
    if number == 19:
        print("regex19")
        regex_string = re.compile(r"^19")
    if number == 20:
        print("regex20")
        regex_string = re.compile(r"^20")
    return regex_string


def remove_numbers(test_action):

    action_text = ""
    if "." in test_action:
        point_pos = test_action.find(".")
        if (point_pos > 0) and (point_pos < 4):  # Take first point after number, not ending point
            action_text = test_action[point_pos + 1:]
            print("action_text", action_text)
    return action_text


def split_steps(nr, rijnr, step, list_r):
    print("split_steps IN parameter nr ", nr)
    print("split_steps IN parameter rijnr ", rijnr)
    print("split_steps IN parameter step ", step)
    print("split_steps IN parameter list_r", list_r)
    write_row_start = rijnr
    write_row_stop = rijnr
    regex_sxx = select_regex_string(nr)
    list_sxx = regex_sxx.findall(step)
    if len(list_sxx) > 0:  # Step01
        stepxx = step
        resultxx = ""
        regex_rxx = select_regex_string(nr)
        match_rxx = 0  # matching pair step - result
        match_rxx_char = 0  # step is present - result has second line
        match_rxx_blank = 0  # steps is present - result has blank line
        match_position = 0
        match_row = 0
        count = 0
        for y in list_r:
            if (match_rxx == 1) and (len(y.strip()) == 0):
                print("<<<skip blank y")
                break
            count += 1
            print("===", count, y)
            list_rxx = regex_rxx.findall(y)
            if len(list_rxx) > 0:  # Matching steps 1
                match_position = count
                match_rxx = 1
                resultxx = y
                flag_debug_matches = 1
                write_to_excel_template(stepxx, resultxx, write_row_start, flag_remove_number, flag_debug_matches)
            else:
                regex_rxx_char = re.compile(r"^[a-zA-Z]")
                list_rxx_char = regex_rxx_char.findall(y)
                if (count == match_position + 1) and (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    step_blank = ""
                    resultxx = y
                    flag_debug_matches = 2
                    write_to_excel_template(step_blank, resultxx, write_row_stop,
                                            flag_remove_number, flag_debug_matches)
                    continue  # stop zoeken naar results
                if (match_rxx_char == 1) and (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    step_blank = ""
                    resultxx = y
                    flag_debug_matches = 3
                    write_to_excel_template(step_blank, resultxx, write_row_stop,
                                            flag_remove_number, flag_debug_matches)
                    continue
                if (match_rxx_blank == 0) and (match_rxx == 0) and (len(y.strip()) == 0):  # result01 = blank
                    match_rxx_blank = 1
                    resultxx = ""
                    flag_debug_matches = 4
                    write_to_excel_template(stepxx, resultxx, write_row_start, flag_remove_number, flag_debug_matches)
                    continue
        if list_r == "":
            resultxx = ""
            flag_debug_matches = 5
            write_to_excel_template(stepxx, resultxx, write_row_start, flag_remove_number, flag_debug_matches)
        elif (match_rxx == 0):
            resultxx = ""
            flag_debug_matches = 6
            write_to_excel_template(stepxx, resultxx, write_row_start, flag_remove_number, flag_debug_matches)
    print("OUT parameter", write_row_stop)
    return write_row_stop


def check_steps_and_results(read_row, read_col, row_start):
    print("check_steps_and_results IN parameter read_row ", read_row)
    print("check_steps_and_results IN parameter read_col ", read_col)
    print("check_steps_and_results IN parameter row_start ", row_start)

    row_stop = row_start
    e = tab.cell(row=read_row, column=read_col)
    if e.value is not None:

        s = tab.cell(row=read_row, column=read_col)
        r = tab.cell(row=read_row, column=read_col + 1)
        if (s.value is not None):  # Steps are present
            test_steps = s.value
            list_steps = test_steps.splitlines()
            if (r.value is not None):  # Results are present
                test_results = r.value
                list_results = test_results.splitlines()
            else:
                test_results = ""
                list_results = ""
#            if len(list_steps) != len(list_results):
            print("++++++++++++++++++++++>", row, len(list_steps), len(list_results))

# Steps Renumbering
            print("------ S T E P S ----------")
            list_steps = renumber_testcase_action(list_steps, list_results)
# Results Renumbering
#            print("------ R E S U L T S ----------")
#            list_results = renumber_testcase_action(list_results)
            cs = 0
            without_number = 0  # steps and numbers without numbers
            for x in list_steps:
                cs += 1
                print(cs, x)
                if "Precondition" in x:
                    precondition = x
                    if "." in x:
                        split_precondition_steps = x.split('.')  # Check first point
                        print("YYYYYYYY split_precondition_steps", split_precondition_steps)
                        print("YYYYYYYY len(split_precondition_steps)", len(split_precondition_steps))
                        if (len(split_precondition_steps) > 0):
                            for i in range(0, len(split_precondition_steps)):
                                print("i", i)
                                print("split_precondition_steps[i] :", split_precondition_steps[i], "YYYYYYYYYYYYY")
                                if "Precondition" in split_precondition_steps[i]:
                                    print("Precondition Catch", split_precondition_steps[i])
                                    x = split_precondition_steps[i].strip()
                                    break
                    write_col = 4  # Steps : read column F(6); Preconditions : write column E(5)
                    print("Precondition ", cs, x.strip())
                    # precondition = tab.cell(row=row_start, column=write_col)
                    # precondition.value = x.strip()
                    worksheet.write(row_start, write_col, x.strip())
                    continue
                if len(x.strip()) == 0:
                    print("Steps Blank", cs, x)
                    continue
#                    for i in range(9):
#                        stepnr = i + 1
                regex_stepnr = re.compile(r"^[0-9][0-9]?")
                list_stepnr = regex_stepnr.findall(x)
                if len(list_stepnr) == 1:
                    int_stepnr = int(list_stepnr[0])
                    print("int_stepnr", type(int_stepnr), int_stepnr)
                    print("split before", row_start)
                    row_start = split_steps(int_stepnr, row_start, x, list_results)
                    print("split after", row_start)
                    row_start += 1
                else:  # Steps and Results without numbers
                    if without_number == 0:
                        steps_start = row_start
                        without_number = 1
                        print("step_start", steps_start)
                    regex_steps = re.compile(r"^[a-zA-Z][a-zA-Z]", re.MULTILINE)
                    list_stepxx = regex_steps.findall(test_steps)
                    print("len(list_stepxx)", len(list_stepxx), list_stepxx)
                    print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                    print("test_steps", len(test_steps), test_steps)
                    print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                    print("test_results", test_results)
                    if (len(list_stepxx) > 0):
                        split_test_steps = test_steps.split('\n')
                        print("XXXXXXXXX split_test_steps", split_test_steps)
                        print("XXXXXXXXX len(split_test_steps)", len(split_test_steps))
                        c = len(list_stepxx) - 1
                        print("c: ", c)
                        for i in range(1, len(split_test_steps)):
                            print("i", i)
                            print("split_test_steps[1] :", split_test_steps[1], "XXXXXXXXXXXX")
                            # print("split", split_test_steps[i])
                            print("list_stepxx[c] :", list_stepxx[c], "XXXXXXXXXXXX")
                            if split_test_steps[i].startswith(list_stepxx[c]):
                                print("Catch", split_test_steps[i])
                                test_steps = split_test_steps[i]
                                test_results = ""
                                break
                        stepxx = test_steps
                        resultxx = test_results
                        print("stepxx", stepxx)
                        # if ("Precondition" not in stepxx):
                        flag_debug_matches = 7
                        write_to_excel_template(stepxx, resultxx, write_row_start,
                                                flag_remove_number, flag_debug_matches)
                        row_start += 1
        else:  # Steps are blank and results are present
            print("STEPS ARE BLANK AND RESULTS ARE PRESENT")
    return row_start


# ------------------MAIN -------------------------


flag_remove_number = 0  # Remove step or result number
flag_debug_matches = 1  # Print debug info function write_step_matches_result
list_col_access = [2, 3, 4, 5, 6]
list_col_copy = [2, 3, 4, 5]
list_col_split = [6, 7]
list_steps = []
list_header1 = ["Test Suite", "Test Cases"]
list_header2 = ["Name", "Details", "Name", "Summary", "Preconditions", "Test Execution Type",
                "Importance", "Steps", "Expected Results", "Step Execution Type", "Requirements"]
write_row_start = 2
write_row_stop = 2
list_TC_topics = ["Install & Uninstall", "First Run & Registration", "Settings",
                  "New Label and Open Label", "Connect & Recognize Printers",
                  "Printers & Consumables", "Text Based Objects", "Graphic Based Objects"]
tabnr = select_excel_tab()
inputfile = "Reviewed Test cases_DCD_Updated_09042018.xlsx"
wb = openpyxl.load_workbook(inputfile)
wb_template = xlsxwriter.Workbook('TC_template.xlsx')
worksheet = wb_template.add_worksheet("TestCases")
# wb = openpyxl.load_workbook('Test cases_DCD_FN.xlsx')
for sheet in wb.sheetnames:
    if sheet == "Totaal reviewed cases":
        continue
    if sheet == "Table of Contents":
        write_header()
        continue  # discard TOC
    if sheet != list_TC_topics[tabnr]:
        continue
    if sheet in list_TC_topics:
        print(sheet, list_TC_topics[tabnr])
        tab = wb[sheet]
        for row in range(2, tab.max_row + 1):
            #  write_row_start += 1
            for col in range(1, tab.max_column + 1):
                if col not in list_col_access:
                    e = ""
                    continue
                e = tab.cell(row=row, column=col)
                if col in list_col_copy:
                    if e.value is not None:
                        if col == 2:  # Traceablity ID : read column B (02); Requirements : write column K(11)
                            write_col = 10
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 3:  # Functionality : read column C (03); Name : write column A(01)
                            write_col = 0
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 4:  # Task : read column D (04); Name : write column C(03)
                            write_col = 2
                            e = tab.cell(row=row, column=col)
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 5:  # Description : read column E (05); Summary : write column D(04)
                            write_col = 3
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 8:  # Priority : read column H (8); Importance : write column G(7)
                            write_col = 6
                            worksheet.write(write_row_start, write_col, e.value)
                if col in list_col_split:
                    print("row_start", write_row_start)
                    write_row_start = check_steps_and_results(row, col, write_row_start)
wb_template.close()  # save the workbook
wb.close()
