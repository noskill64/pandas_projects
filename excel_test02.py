import openpyxl
import re
import xlsxwriter


def split_steps(nr, rijnr, step, list_r):
    print("IN parameter nr ", nr)
    print("IN parameter rijnr ", rijnr)
    print("IN parameter step ", step)
    print("IN parameter list_r", list_r)
    if nr == 1:
        print("regex1")
        write_row_start = rijnr
        write_row_stop = rijnr
        regex_sxx = re.compile(r"^1")
    if nr == 2:
        print("regex2")
        regex_sxx = re.compile(r"^2")
    if nr == 3:
        print("regex3")
        regex_sxx = re.compile(r"^3")
    if nr == 4:
        print("regex4")
        regex_sxx = re.compile(r"^4")
    if nr == 5:
        regex_sxx = re.compile(r"^5")
    if nr == 6:
        regex_sxx = re.compile(r"^6")
    if nr == 7:
        regex_sxx = re.compile(r"^7")
    if nr == 8:
        regex_sxx = re.compile(r"^8")
    if nr == 9:
        regex_sxx = re.compile(r"^9")

    list_sxx = regex_sxx.findall(step)
    if len(list_sxx) > 0:  # Step01
        stepxx = step
        resultxx = ""
        for y in list_r:
            if nr == 1:
                print(y)
                regex_rxx = re.compile(r"^1")
            if nr == 2:
                regex_rxx = re.compile(r"^2")
            if nr == 3:
                regex_rxx = re.compile(r"^3")
            if nr == 4:
                regex_rxx = re.compile(r"^4")
            if nr == 5:
                regex_rxx = re.compile(r"^5")
            if nr == 6:
                regex_rxx = re.compile(r"^6")
            if nr == 7:
                regex_rxx = re.compile(r"^7")
            if nr == 8:
                regex_rxx = re.compile(r"^8")
            if nr == 9:
                regex_rxx = re.compile(r"^9")
            list_rxx = regex_rxx.findall(y)
            if len(list_rxx) > 0:  # Matching steps 1
                match_rxx = 1
                resultxx = y
                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                worksheet.write(write_row_start, write_col, stepxx)
                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                worksheet.write(write_row_start, write_col, resultxx)
                print("result ", nr, stepxx, resultxx)
                continue  # check volgende lijn op characters
            else:
                regex_rxx_char = re.compile(r"^[a-zA-Z]")
                list_rxx_char = regex_rxx_char.findall(y)
                if (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    resultxx = y
                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                    worksheet.write(write_row_stop, write_col, stepxx)
                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                    worksheet.write(write_row_stop, write_col, resultxx)  # schrijf lijn + 1
                    print("result", nr, "NOT NULL", stepxx, resultxx)
                    break  # stop zoeken naar results
                if (match_rxx == 0) and (len(y.strip()) == 0):  # result01 = blank
                    match_rxx_blank = 1
                    resultxx = ""
                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                    worksheet.write(write_row_start, write_col, stepxx)
                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                    worksheet.write(write_row_start, write_col, resultxx)
                    print("result", nr, " BLANK", stepxx, resultxx)
                    break  # stop zoeken naar results
    print("OUT parameter", write_row_stop)
    return write_row_stop


def check_steps_and_results(read_row, read_col, row_start):

    row_stop = row_start
    e = tab.cell(row=read_row, column=read_col)
    if e.value is not None:

        s = tab.cell(row=read_row, column=read_col)
        r = tab.cell(row=read_row, column=read_col + 1)
        if (s.value is not None) and (r.value is not None):
            test_steps = s.value
            test_results = r.value
            list_steps = test_steps.splitlines()
            list_results = test_results.splitlines()
            if len(list_steps) != len(list_results):
                print("++++++++++++++++++++++>", row, len(list_steps), len(list_results))
                cs = 0
                step01 = step02 = step03 = step04 = step05 = ""
                step06 = step07 = step08 = step09 = ""
                for x in list_steps:
                    cs += 1
                    print(cs, x)
                    if "Steps Precondition" in x:
                        precondition = x
                        write_col = 4  # Steps : read column F(6); Preconditions : write column E(5)
                        worksheet.write(row_start, write_col, precondition)
                        print("Steps Precondition ", cs, precondition)
                        continue
                    if len(x.strip()) == 0:
                        print("Steps Blank", cs, x)
                        continue
                    for i in range(9):
                        stepnr = i + 1
                        row_start = split_steps(stepnr, row_start, x, list_results)
                        row_start += 1


# ------------------MAIN -------------------------
workbook = xlsxwriter.Workbook('test case template.xlsx')
worksheet = workbook.add_worksheet("TestCases")


list_col_access = [2, 4, 5, 6]
list_col_copy = [2, 4, 5]
list_col_split = [6, 7]
list_steps = []
write_row_start = 2
write_row_stop = 2


# wb = openpyxl.load_workbook('Test cases_DCD_Updated_03212018.xlsx')
wb = openpyxl.load_workbook('Test cases_DCD_FN.xlsx')
for sheet in wb.sheetnames:
    if sheet == "Table of Contents":
        continue  # discard TOC
#    print(sheet)
    if sheet == "Install & Uninstall":
        list_steps = []
        write_row_start = 2
        write_row_stop = 2
        tab = wb[sheet]
        for row in range(2, tab.max_row + 1):
            write_row_start += 1
            for col in range(1, tab.max_column + 1):
                if col not in list_col_access:
                    e = ""
                    continue
                e = tab.cell(row=row, column=col)
                if col in list_col_copy:
                    if e.value is not None:
                        if col == 2:  # Traceablity ID : read column B (2); Requirements : write column K(11)
                            write_col = 10
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 4:  # Task : read column D (4); Name : write column C(3)
                            write_col = 2
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 5:  # Description : read column E (5); Summary : write column D(4)
                            write_col = 3
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 8:  # Priority : read column H (8); Importance : write column G(7)
                            write_col = 6
                            worksheet.write(write_row_start, write_col, e.value)
                if col in list_col_split:
                    check_steps_and_results(row, col, write_row_start)
