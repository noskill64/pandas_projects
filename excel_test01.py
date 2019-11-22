import openpyxl
import re
import xlsxwriter


workbook = xlsxwriter.Workbook('test case template.xlsx')
worksheet = workbook.add_worksheet("TestCases")


list_col_access = [2, 4, 5, 6]
list_col_copy = [2, 4, 5]
list_col_split = [6, 7]
list_steps = []
write_row_start = 2
write_row_stop = 2


wb = openpyxl.load_workbook('Test cases_DCD_Updated_03212018.xlsx')
for sheet in wb.sheetnames:
    if sheet == "Table of Contents":
        continue  # discard TOC
#    print(sheet)
    if sheet == "Install & Uninstall":
        list_steps = []
        write_row_start = 2
        write_row_stop = 2
        tab = wb[sheet]
        for row in range(2, tab.max_row + 1):
            write_row_start += 1
            for col in range(1, tab.max_column + 1):
                if col not in list_col_access:
                    e = ""
                    continue
                e = tab.cell(row=row, column=col)
                if col in list_col_copy:
                    if e.value is not None:
                        if col == 2:  # Traceablity ID : read column B (2); Requirements : write column K(11)
                            write_col = 10
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 4:  # Task : read column D (4); Name : write column C(3)
                            write_col = 2
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 5:  # Description : read column E (5); Summary : write column D(4)
                            write_col = 3
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 8:  # Priority : read column H (8); Importance : write column G(7)
                            write_col = 6
                            worksheet.write(write_row_start, write_col, e.value)
                if col in list_col_split:
                    write_row_stop = write_row_start
                    if e.value is not None:

                        s = tab.cell(row=row, column=col)
                        r = tab.cell(row=row, column=col + 1)
                        if (s.value is not None) and (r.value is not None):
                            test_steps = s.value
                            test_results = r.value
                            list_steps = test_steps.splitlines()
                            list_results = test_results.splitlines()
                            if len(list_steps) != len(list_results):
                                print("++++++++++++++++++++++>", row, len(list_steps), len(list_results))
                                cs = 0
                                step01 = step02 = step03 = step04 = step05 = ""
                                step06 = step07 = step08 = step09 = ""
                                for x in list_steps:
                                    match_r01 = match_r01_char = match_r01_blank = 0
                                    match_r02 = match_r02_char = match_r02_blank = 0
                                    match_r03 = match_r03_char = match_r03_blank = 0
                                    match_r04 = match_r04_char = match_r04_blank = 0
                                    match_r05 = match_r05_char = match_r05_blank = 0
                                    match_r06 = match_r06_char = match_r06_blank = 0
                                    match_r07 = match_r07_char = match_r07_blank = 0
                                    match_r08 = match_r08_char = match_r08_blank = 0
                                    match_r09 = match_r09_char = match_r09_blank = 0
                                    cs += 1
                                    print(cs, x)
                                    if "Steps Precondition" in x:
                                        precondition = x
                                        write_col = 4  # Steps : read column F(6); Preconditions : write column E(5)
                                        worksheet.write(write_row_start, write_col, precondition)
                                        print("Steps Precondition ", cs, precondition)
                                        continue
                                    if len(x.strip()) == 0:
                                        print("Steps Blank", cs, x)
                                        continue

                                    regex_s01 = re.compile(r"^1")
                                    list_s01 = regex_s01.findall(x)
                                    if len(list_s01) > 0:  # Step01
                                        step01 = x
                                        result01 = ""
                                        for y in list_results:
                                            regex_r01 = re.compile(r"^1")
                                            list_r01 = regex_r01.findall(y)
                                            if len(list_r01) > 0:  # Matching steps 1
                                                match_r01 = 1
                                                result01 = y
                                                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                worksheet.write(write_row_start, write_col, step01)
                                                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                worksheet.write(write_row_start, write_col, result01)
                                                print("result01", step01, result01)
                                                continue  # check volgende lijn op characters
                                            else:
                                                regex_r01_char = re.compile(r"^[a-zA-Z]")
                                                list_r01_char = regex_r01_char.findall(y)
                                                if (match_r01 == 1) and (len(list_r01_char > 0)):
                                                    write_row_stop += 1
                                                    match_r01_char = 1
                                                    result01 = y
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_start, write_col, step01)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col,
                                                                    result01)  # schrijf lijn + 1
                                                    print("result01 NOT NULL", step01, result01)
                                                    break  # stop zoeken naar results
                                                if (match_r01 == 0) and (len(y.strip()) == 0):  # result01 = blank
                                                    match_r01_blank = 1
                                                    result01 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_start, write_col, step01)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_start, write_col, result01)
                                                    print("result01 BLANK", step01, result01)
                                                    break  # stop zoeken naar results
                                    if (match_r01 == 1) or (match_r01_char == 1) or (match_r01_blank == 1):
                                        continue  # goto step02
                                    regex_s02 = re.compile(r"^2")
                                    list_s02 = regex_s02.findall(x)
                                    if len(list_s02) > 0:  # Step02
                                        step02 = x
                                        result02 = ""
                                        for y in list_results:
                                            if "2." in y:
                                                result02 = y
                                                write_row_stop += 1
                                                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                worksheet.write(write_row_stop, write_col, step02)
                                                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                worksheet.write(write_row_stop, write_col, result02)
                                                print("result02", step02, result02)
                                                continue
                                            else:
                                                if len(y.strip()) == 0:
                                                    result02 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_stop, write_col, step02)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col, result02)
                                                    print("result02 BLANK", step02, result02)
                                                    continue
                                                if len(y.strip()) != 0:
                                                    result02 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_stop, write_col, step02)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col, result02)
                                                    print("result02 NOT NULL", step02, result02)
                                                    continue

                                    if "3." in x:
                                        step03 = x
                                        result03 = ""
                                        for y in list_test_results:
                                            if "3." in y:
                                                result03 = y
                                                write_row_stop += 1
                                                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                worksheet.write(write_row_stop, write_col, step03)
                                                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                worksheet.write(write_row_stop, write_col, result03)
                                                print("result03", step03, result03)
                                                continue
                                            else:
                                                if len(y.strip()) == 0:
                                                    result03 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_stop, write_col, step03)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col, result03)
                                                    print("result03 BLANK", step03, result03)
                                                    continue
                                                if len(y.strip()) != 0:
                                                    result03 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_stop, write_col, step03)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col, result03)
                                                    print("result03 NOT NULL", step03, result03)
                                                    continue

                                    if ("4." in x) or ("4a." in x) or ("4b." in x):
                                        step04 = x
                                        result04 = ""
                                        for y in list_test_results:
                                            if ("4." in y) or ("4a." in y) or ("4b." in y):
                                                result04 = y
                                                write_row_stop += 1
                                                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                worksheet.write(write_row_stop, write_col, step04)
                                                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                worksheet.write(write_row_stop, write_col, result04)
                                                print("result04", step04, result04)
                                                continue
                                            else:
                                                if len(y.strip()) == 0:
                                                    result04 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_stop, write_col, step04)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col, result04)
                                                    print("result04 BLANK", step04, result04)
                                                    continue

                                    if "5." in x:
                                        step05 = x
                                        result05 = ""
                                        for y in list_test_results:
                                            if "5." in y:
                                                result05 = y
                                                write_row_stop += 1
                                                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                worksheet.write(write_row_stop, write_col, step05)
                                                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                worksheet.write(write_row_stop, write_col, result05)
                                                print("result05", step05, result05)
                                                continue
                                            else:
                                                if len(y.strip()) == 0:
                                                    result05 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_stop, write_col, step05)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col, result05)
                                                    print("result05 BLANK", step05, result05)
                                                    continue

                                    if "6." in x:
                                        step06 = x
                                        result06 = ""
                                        for y in list_test_results:
                                            if "6." in y:
                                                result06 = y
                                                write_row_stop += 1
                                                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                worksheet.write(write_row_stop, write_col, step06)
                                                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                worksheet.write(write_row_stop, write_col, result06)
                                                print("result06", step06, result06)
                                                continue
                                            else:
                                                if len(y.strip()) == 0:
                                                    result06 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_stop, write_col, step06)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col, result06)
                                                    print("result06 BLANK", step06, result06)
                                                    continue

                                    if "7." in x:
                                        step07 = x
                                        result07 = ""
                                        for y in list_test_results:
                                            if "7." in y:
                                                result07 = y
                                                write_row_stop += 1
                                                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                worksheet.write(write_row_stop, write_col, step07)
                                                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                worksheet.write(write_row_stop, write_col, result07)
                                                print("result07", step07, result07)
                                                continue
                                            else:
                                                if len(y.strip()) == 0:
                                                    result07 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_stop, write_col, step07)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col, result07)
                                                    print("result07 BLANK", step07, result07)
                                                    continue

                                    if "8." in x:
                                        step08 = x
                                        result08 = ""
                                        for y in list_test_results:
                                            if "8." in y:
                                                result08 = y
                                                write_row_stop += 1
                                                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                worksheet.write(write_row_stop, write_col, step08)
                                                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                worksheet.write(write_row_stop, write_col, result08)
                                                print("result08", step08, result08)
                                                continue
                                            else:
                                                if len(y.strip()) == 0:
                                                    result08 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_stop, write_col, step08)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col, result08)
                                                    print("result08 BLANK", step08, result08)
                                                    continue

                                    if "9." in x:
                                        step09 = x
                                        result09 = ""
                                        for y in list_test_results:
                                            if "9." in y:
                                                result09 = y
                                                write_row_stop += 1
                                                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                worksheet.write(write_row_stop, write_col, step09)
                                                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                worksheet.write(write_row_stop, write_col, result09)
                                                print("result09 ", step09, result09)
                                                continue
                                            else:
                                                if len(y.strip()) == 0:
                                                    result09 = ""
                                                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                                                    worksheet.write(write_row_stop, write_col, step09)
                                                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                                                    worksheet.write(write_row_stop, write_col, result09)
                                                    print("result09 BLANK", step09, result09)

                            write_row_start = write_row_stop
