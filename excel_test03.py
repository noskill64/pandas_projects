import openpyxl
import re
import xlsxwriter


def split_steps(nr, rijnr, step, list_r):
    print("split_steps IN parameter nr ", nr)
    print("split_steps IN parameter rijnr ", rijnr)
    print("split_steps IN parameter step ", step)
    print("split_steps IN parameter list_r", list_r)
    write_row_start = rijnr
    write_row_stop = rijnr
    if nr == 1:
        print("steps_regex1")
        regex_sxx = re.compile(r"^1")
    if nr == 2:
        print("steps_regex2")
        regex_sxx = re.compile(r"^2")
    if nr == 3:
        print("steps_regex3")
        regex_sxx = re.compile(r"^3")
    if nr == 4:
        print("steps_regex4")
        regex_sxx = re.compile(r"^4")
    if nr == 5:
        print("steps_regex5")
        regex_sxx = re.compile(r"^5")
    if nr == 6:
        print("steps_regex6")
        regex_sxx = re.compile(r"^6")
    if nr == 7:
        print("steps_regex7")
        regex_sxx = re.compile(r"^7")
    if nr == 8:
        print("steps_regex8")
        regex_sxx = re.compile(r"^8")
    if nr == 9:
        print("steps_regex9")
        regex_sxx = re.compile(r"^9")

    list_sxx = regex_sxx.findall(step)
    if len(list_sxx) > 0:  # Step01
        stepxx = step
        resultxx = ""
        if nr == 1:
            print("results_regex1")
            regex_rxx = re.compile(r"^1")
        if nr == 2:
            print("results_regex2")
            regex_rxx = re.compile(r"^2")
        if nr == 3:
            print("results_regex3")
            regex_rxx = re.compile(r"^3")
        if nr == 4:
            print("results_regex4")
            regex_rxx = re.compile(r"^4")
        if nr == 5:
            print("results_regex5")
            regex_rxx = re.compile(r"^5")
        if nr == 6:
            print("results_regex6")
            regex_rxx = re.compile(r"^6")
        if nr == 7:
            print("results_regex7")
            regex_rxx = re.compile(r"^7")
        if nr == 8:
            print("results_regex8")
            regex_rxx = re.compile(r"^8")
        if nr == 9:
            print("results_regex9")
            regex_rxx = re.compile(r"^9")
        match_rxx = 0  # matching pair step - result
        match_rxx_char = 0  # step is present - result has second line
        match_rxx_blank = 0  # steps is present - result has blank line
        match_position = 0
        count = 0
        for y in list_r:
            #  if (len(y.strip()) == 0):
            #  print("<<<skip blank y")
            #  continue
            count += 1
            print("===", count, y)
            list_rxx = regex_rxx.findall(y)
            if len(list_rxx) > 0:  # Matching steps 1
                match_position = count
                match_rxx = 1
                resultxx = y
                write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                worksheet.write(write_row_start, write_col, stepxx)
                write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                worksheet.write(write_row_start, write_col, resultxx)
                print("result ", nr, stepxx, resultxx, write_row_start)
            else:
                regex_rxx_char = re.compile(r"^[a-zA-Z]")
                list_rxx_char = regex_rxx_char.findall(y)
                if (count == match_position + 1) and (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    step_blank = ""
                    resultxx = y
                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                    worksheet.write(write_row_stop, write_col, step_blank)
                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                    worksheet.write(write_row_stop, write_col, resultxx)  # schrijf lijn + 1
                    print("result", nr, "NOT NULL", step_blank, resultxx, write_row_stop)
                    continue  # stop zoeken naar results
                if (match_rxx_char == 1) and (match_rxx == 1) and (len(list_rxx_char) > 0):
                    write_row_stop += 1
                    match_rxx_char = 1
                    step_blank = ""
                    resultxx = y
                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                    worksheet.write(write_row_stop, write_col, step_blank)
                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                    worksheet.write(write_row_stop, write_col, resultxx)  # schrijf lijn + 1
                    print("NEXT ABC RESULTS", nr, step_blank, resultxx, write_row_stop)
                    continue
                if (match_rxx_blank == 0) and (match_rxx == 0) and (len(y.strip()) == 0):  # result01 = blank
                    match_rxx_blank = 1
                    resultxx = ""
                    write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                    worksheet.write(write_row_start, write_col, stepxx)
                    write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                    worksheet.write(write_row_start, write_col, resultxx)
                    print("result", nr, " BLANK", stepxx, resultxx, write_row_start)
                    continue
        if list_r == "":
            resultxx = ""
            write_col = 7  # Steps : read column F(6); Steps : write column H(8)
            worksheet.write(write_row_start, write_col, stepxx)
            write_col = 8  # Steps : read column F(6); Steps : write column I(9)
            worksheet.write(write_row_start, write_col, resultxx)
            print("MISSING RESULTS", nr, " BLANK", stepxx, resultxx, write_row_start)
        elif (match_rxx == 0):
            resultxx = ""
            write_col = 7  # Steps : read column F(6); Steps : write column H(8)
            worksheet.write(write_row_start, write_col, stepxx)
            write_col = 8  # Steps : read column F(6); Steps : write column I(9)
            worksheet.write(write_row_start, write_col, resultxx)
            print("WHAT THE FUCK", list_r, "$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("OUT parameter", write_row_stop)
    return write_row_stop


def check_steps_and_results(read_row, read_col, row_start):
    print("check_steps_and_results IN parameter read_row ", read_row)
    print("check_steps_and_results IN parameter read_col ", read_col)
    print("check_steps_and_results IN parameter row_start ", row_start)

    row_stop = row_start
    e = tab.cell(row=read_row, column=read_col)
    if e.value is not None:

        s = tab.cell(row=read_row, column=read_col)
        r = tab.cell(row=read_row, column=read_col + 1)
        if (s.value is not None):  # Steps are present
            test_steps = s.value
            list_steps = test_steps.splitlines()
            if (r.value is not None):  # Results are present
                test_results = r.value
                list_results = test_results.splitlines()
            else:
                test_results = ""
                list_results = ""
#            if len(list_steps) != len(list_results):
            print("++++++++++++++++++++++>", row, len(list_steps), len(list_results))
            cs = 0
            step01 = step02 = step03 = step04 = step05 = ""
            step06 = step07 = step08 = step09 = ""
            for x in list_steps:
                cs += 1
                print(cs, x)
                if "Precondition" in x:
                    precondition = x
                    write_col = 4  # Steps : read column F(6); Preconditions : write column E(5)
                    print("Precondition ", cs, x)
                    precondition = tab.cell(row=row_start, column=write_col)
                    precondition.value = x
                    worksheet.write(row_start, write_col, precondition.value)
                    continue
                if len(x.strip()) == 0:
                    print("Steps Blank", cs, x)
                    continue
#                    for i in range(9):
#                        stepnr = i + 1
                regex_stepnr = re.compile(r"^[0-9]")
                list_stepnr = regex_stepnr.findall(x)
                if len(list_stepnr) == 1:
                    int_stepnr = int(list_stepnr[0])
                    print("int_stepnr", type(int_stepnr), int_stepnr)
                    print("split before", row_start)
                    row_start = split_steps(int_stepnr, row_start, x, list_results)
                    print("split after", row_start)
                    row_start += 1
                else:  # Steps and Results without numbers
                    regex_steps = re.compile(r"^[a-zA-Z]")
                    list_stepxx = regex_steps.findall(test_steps)
                    print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                    print("len(list_stepxx)", len(list_stepxx))
                    print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                    print("test_steps", test_steps)
                    print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                    print("test_results", test_results)
                    if len(list_stepxx) > 0:
                        stepxx = test_steps
                        resultxx = test_results
                        write_col = 7  # Steps : read column F(6); Steps : write column H(8)
                        worksheet.write(row_start, write_col, stepxx)
                        write_col = 8  # Steps : read column F(6); Steps : write column I(9)
                        worksheet.write(row_start, write_col, resultxx)
                    print("STEPS WITHOUT NUMBERS ???? ", "1", stepxx, resultxx, row_start)
                    row_start += 1
        else:  # Steps are blank and results are present
            print("STEPS ARE BLANK AND RESULTS ARE PRESENT")
    return row_start


# ------------------MAIN -------------------------


list_col_access = [2, 4, 5, 6]
list_col_copy = [2, 4, 5]
list_col_split = [6, 7]
list_steps = []
list_header1 = ["Test Suite", "Test Cases"]
list_header2 = ["Name", "Details", "Name", "Summary", "Preconditions", "Test Execution Type",
                "Importance", "Steps", "Expected Results", "Step Execution Type", "Requirements"]
write_row_start = 2
write_row_stop = 2
list_TC_topics = ["Install & Uninstall", "First Run & Registration", "Settings"]


wb = openpyxl.load_workbook('Test cases_DCD_Updated_03212018.xlsx')
wb_template = xlsxwriter.Workbook('TC_template.xlsx')
worksheet = wb_template.add_worksheet("TestCases")
# wb = openpyxl.load_workbook('Test cases_DCD_FN.xlsx')
for sheet in wb.sheetnames:
    if sheet == "Table of Contents":
        #  Print header 1
        tab = wb[sheet]
        header1_row = 0
        header1_col1 = 0
        header1_col2 = 2
        print("Header1", header1_row, header1_col1, list_header1[0])  # Header 1 Column Test Suite
        worksheet.write(header1_row, header1_col1, list_header1[0])
        print("Header1", header1_row, header1_col2, list_header1[1])  # Header 1 Column Test Cases
        worksheet.write(header1_row, header1_col2, list_header1[1])
        #  Print header 2
        header2_row = 1
        header2_max_col = 11
        for header2_col in range(header2_max_col):
            #   header2 = tab.cell(row=header2_row, column=header2_col)
            #   header2.value = list_header2[header2_col - 1]
            print("Header2", header2_row, header2_col, list_header2[header2_col])
            worksheet.write(header2_row, header2_col, list_header2[header2_col])
        continue  # discard TOC
    if sheet in list_TC_topics:
        print("START OF ", sheet, "+++++++++++++++++++++++++++++++++++++")
        list_steps = []
        tab = wb[sheet]
        for row in range(2, tab.max_row + 1):
            #  write_row_start += 1
            for col in range(1, tab.max_column + 1):
                if col not in list_col_access:
                    e = ""
                    continue
                e = tab.cell(row=row, column=col)
                if col in list_col_copy:
                    if e.value is not None:
                        if col == 2:  # Traceablity ID : read column B (2); Requirements : write column K(11)
                            write_col = 10
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 4:  # Task : read column D (4); Name : write column C(3)
                            write_col = 2
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 5:  # Description : read column E (5); Summary : write column D(4)
                            write_col = 3
                            worksheet.write(write_row_start, write_col, e.value)
                        if col == 8:  # Priority : read column H (8); Importance : write column G(7)
                            write_col = 6
                            worksheet.write(write_row_start, write_col, e.value)
                if col in list_col_split:
                    print("main_write_row_start", write_row_start)
                    write_row_start = check_steps_and_results(row, col, write_row_start)
wb_template.close()  # save the workbook
wb.close()
